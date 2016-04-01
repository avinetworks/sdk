import json
import copy
import numbers
import logging
import os

from openpyxl import Workbook

LOG = logging.getLogger("converter-log")
row_count = 1
ws = None


def get_port_by_protocol(protocol):
    """
    Instead of default ports for protocols F5 config has protocol in
    destination value for Avi object need to conver it to port number
    :param protocol: protocol name
    :return: integer value for protocol
    """
    port = 80
    if protocol == "https":
        port = 443
    elif protocol == "ftp":
        port = 21
    elif protocol == "smtp":
        port = 25
    elif protocol == "snmp":
        port = 161
    elif protocol == "telnet":
        port = 23
    elif protocol == "snmp-trap":
        port = 162
    elif protocol == "ssh":
        port = 22
    return port


def convert_servers_config(servers_config):
    """
    Converts the config of servers in the pool
    :param servers_config: F5 servers config for particular pool
    :return: List of Avi server configs
    """
    server_list = []
    skipped_list = []
    supported_attributes = ['session']
    for server_name in servers_config.keys():
        skipped = None
        server = servers_config[server_name]
        parts = server_name.split(':')
        ip_addr = parts[0]
        port = parts[1] if len(parts) == 2 else 80
        if not port.isdigit():
            port = get_port_by_protocol(port)
        enabled = True
        state = 'enabled'
        if server:
            state = server.get("session", 'enabled')
            skipped = [key for key in server.keys()
                       if key not in supported_attributes]
        if state == "user disabled":
            enabled = False
        server_list.append({
            'ip': {
                'addr': ip_addr,
                'type': 'V4'
            },
            'port': port,
            'enabled': enabled
        })

        if skipped:
            skipped_list.append({server_name: skipped})
    return server_list, skipped_list


def get_avi_pool_down_action(action):
    """
    Maps Pool down action from F5 config to Avi Config
    :param action: F5 action string
    :return: Avi action String
    """
    if action == "reset":
        return "POOL_DOWN_ACTION_CLOSE_CONN"
    if action == "reselect":
        return "POOL_DOWN_ACTION_CLOSE_CONN"
    else:
        return "POOL_DOWN_ACTION_CLOSE_CONN"


def get_avi_lb_algorithm(f5_algorithm):
    avi_algorithm = None
    if not f5_algorithm or f5_algorithm in ["ratio", "member ratio"]:
        avi_algorithm = "LB_ALGORITHM_ROUND_ROBIN"
    elif f5_algorithm in ["member least conn", "least conn",
                          "weighted least conn member", "l3 addr"
                          "weighted least conn node addr", "least sessions"]:
        avi_algorithm = "LB_ALGORITHM_LEAST_CONNECTIONS"
    elif f5_algorithm in ["fastest", "fastest app resp"]:
        avi_algorithm = "LB_ALGORITHM_FASTEST_RESPONSE"
    elif f5_algorithm in ["dynamic ratio", "member observed", "predictive",
                          "member predictive", "observed",
                          "member dynamic ratio"]:
        avi_algorithm = "LB_ALGORITHM_LEAST_LOAD"
    return avi_algorithm


def convert_pool_config(pool_config, tenant, monitor_config_list):
    """
    Convert list of pools from F5 config to Avi config
    :param pool_config: F5 pool config
    :param tenant: name of tenant for default objects
    :param monitor_config_list: Avi monitor config list
    :return: List of pools converted to Avi configuration
    """
    pool_list = []
    supported_attr = ['members', 'monitor', 'action on svcdown', 'lb method']
    for pool_name in pool_config.keys():
        skipped = []
        f5_pool = pool_config[pool_name]
        if not f5_pool:
            LOG.debug("Empty pool skipped for conversion :"+pool_name)
            add_status_row('pool', None, pool_name, 'skipped', None, None)
            continue
        servers, member_skipped_config = convert_servers_config(
            f5_pool.get("members", {}))
        sd_action = f5_pool.get("action on svcdown", "")
        pd_action = get_avi_pool_down_action(sd_action)
        lb_method = f5_pool.get("lb method", None)
        lb_algorithm = get_avi_lb_algorithm(lb_method)
        pool_obj = {
                "name": pool_name,
                "servers": servers,
                "pd_action_type": pd_action,
                "lb_algorithm": lb_algorithm
            }
        monitor_names = f5_pool.get("monitor", None)
        default_monitors = {"http": "System-HTTP", "https": "System-HTTPS",
                            "tcp": "System-TCP", "icmp": "System-Ping",
                            "udp": "System-UDP", "gateway_icmp": "System-Ping"}
        is_ssl = False
        skipped_monitors = []
        if monitor_names:
            monitors = monitor_names.split(" ")
            monitor_refs = []
            for monitor in monitors:
                if monitor in ["and", "all", "min", "of"] or monitor.isdigit():
                    continue
                monitor_obj = [obj for obj in monitor_config_list
                               if obj["name"] == monitor]
                if monitor_obj:
                    monitor_refs.append(monitor_obj[0]["name"])
                    if monitor_obj[0]["type"] == "HEALTH_MONITOR_HTTPS":
                        is_ssl = True
                elif monitor in default_monitors.keys():
                    if monitor == "https":
                        is_ssl = True
                    monitor = tenant+":"+default_monitors[monitor]
                    monitor_refs.append(monitor)
                else:
                    LOG.warning("Skiping non supported monitor: %s for pool %s"
                                % (monitor, pool_name))
                    skipped_monitors.append(monitor)
            pool_obj["health_monitor_refs"] = monitor_refs
        if is_ssl:
            pool_obj["ssl_profile_ref"] = "admin:System-Standard"
        pool_list.append(pool_obj)
        skipped_attr = [key for key in f5_pool.keys() if
                        key not in supported_attr]
        if skipped_attr:
            skipped.append(skipped_attr)
        if member_skipped_config:
            skipped.append(member_skipped_config)
        if skipped_monitors:
            skipped.append({"monitors": skipped_monitors})
        if skipped:
            add_status_row('pool', None, pool_name, 'partial',
                           skipped, pool_obj)
        else:
            add_status_row('pool', None, pool_name, 'successful',
                           skipped, pool_obj)
    return pool_list


def get_defaults(f5_monitor, monitor_config):
    """
    Monitor can have inheritance used by attribute defaults-from in F5
    configuration this method recursively gets all the attributes from the
    default objects and forms complete object
    :param f5_monitor: F5 monitor object
    :param monitor_config: List of F5 monitor configs
    :return:
    """
    parent_name = f5_monitor.get("defaults from", None)
    if parent_name:
        parent_monitor = monitor_config.get(parent_name, None)
        if parent_monitor:
            parent_monitor = get_defaults(parent_monitor, monitor_config)
            parent_monitor = copy.deepcopy(parent_monitor)
            parent_monitor.update(f5_monitor)
            f5_monitor = parent_monitor
        else:
            f5_monitor["type"] = parent_name
    return f5_monitor


def convert_monitor_entity(name, f5_monitor):
    """
    Conversion of single F5 monitor object to Avi health monitor object
    :param name: name of health monitor
    :param f5_monitor: F5 monitor config object
    :return: Avi monitor config object
    """
    supported_attributes = ["timeout", "interval", "time until up",
                            "description", "type", "defaults from"]
    skipped = [key for key in f5_monitor.keys()
               if key not in supported_attributes]
    timeout = int(f5_monitor.get("timeout", 16))
    interval = int(f5_monitor.get("interval", 5))
    time_until_up = int(f5_monitor.get("time until up", 1))
    successful_checks = int(timeout/interval)
    if time_until_up > 0:
        failed_checks = int(time_until_up/interval)
        failed_checks = 1 if failed_checks == 0 else failed_checks
    else:
        failed_checks = 1
    description = f5_monitor.get("description", None)
    monitor_dict = dict()
    monitor_dict["name"] = name
    monitor_dict["receive_timeout"] = interval-1
    monitor_dict["failed_checks"] = failed_checks
    monitor_dict["send_interval"] = interval
    monitor_dict["successful_checks"] = successful_checks
    if description:
        monitor_dict["description"] = description

    if f5_monitor["type"] == "http":
        monitor_dict["type"] = "HEALTH_MONITOR_HTTP"
        monitor_dict["http_monitor"] = {
            "http_request": "HEAD / HTTP/1.0",
            "http_response_code": [
                {"code": "HTTP_2XX"}, {"code": "HTTP_3XX"}
            ]}
    elif f5_monitor["type"] == "https":
        monitor_dict["type"] = "HEALTH_MONITOR_HTTPS"
        monitor_dict["https_monitor"] = {
            "http_request": "HEAD / HTTP/1.0",
            "http_response_code": [
                {"code": "HTTP_2XX"}, {"code": "HTTP_3XX"}
            ]}
    elif f5_monitor["type"] == "tcp":
        tcp_attr = ["dest", "send", "recv"]
        skipped = [key for key in skipped if key not in tcp_attr]
        destination = f5_monitor.get("dest", "*:*")
        dest_str = destination.split(":")
        if len(dest_str) > 1 and isinstance(dest_str[1], numbers.Integral):
            monitor_dict["monitor_port"] = dest_str[1]
        monitor_dict["type"] = "HEALTH_MONITOR_TCP"
        request = f5_monitor.get("send", None)
        response = f5_monitor.get("recv", None)
        if request or response:
            tcp_monitor = {"tcp_request": request, "tcp_response": response}
            monitor_dict["tcp_monitor"] = tcp_monitor
    elif f5_monitor["type"] == "udp":
        udp_attr = ["dest", "send", "recv"]
        skipped = [key for key in skipped if key not in udp_attr]
        destination = f5_monitor.get("dest", "*:*")
        dest_str = destination.split(":")
        if len(dest_str) > 1 and isinstance(dest_str[1], numbers.Integral):
            monitor_dict["monitor_port"] = dest_str[1]
        monitor_dict["type"] = "HEALTH_MONITOR_UDP"
        request = f5_monitor.get("send", None)
        response = f5_monitor.get("recv", None)
        if request or response:
            tcp_monitor = {"udp_request": request, "udp_response": response}
            monitor_dict["tcp_monitor"] = tcp_monitor
    elif f5_monitor["type"] in ["gateway_icmp", "icmp"]:
        monitor_dict["type"] = "HEALTH_MONITOR_PING"
    elif f5_monitor["type"] == "external":
        ext_attr = ["run", "args", "user-defined"]
        skipped = [key for key in skipped if key not in ext_attr]
        monitor_dict["type"] = "HEALTH_MONITOR_EXTERNAL"
        ext_monitor = {
            "command_code": f5_monitor.get("run", None),
            "command_parameters": f5_monitor.get("args", None),
            "command_variables": f5_monitor.get("user-defined", None)
        }
        monitor_dict["external_monitor"] = ext_monitor
    return monitor_dict, skipped


def convert_monitor_config(monitor_config):
    """
    Convert F5 monitor config dict to Avi health monitor config list
    :param monitor_config: F5 monitor config dict
    :return: List of Avi health monitor objects
    """
    monitor_list = []
    supported_types = ["http", "https", "dns", "external", "tcp", "udp",
                       "gateway_icmp", "icmp"]
    for key in monitor_config.keys():
        f5_monitor = monitor_config[key]
        if not f5_monitor:
            add_status_row('monitor', '', key, 'skipped', None, None)
            continue
        f5_monitor = get_defaults(f5_monitor, monitor_config)
        if f5_monitor["type"] not in supported_types:
            LOG.debug("Monitor type not supported by Avi : "+key)
            add_status_row('monitor', f5_monitor["type"], key, 'skipped',
                           None, None)
            continue
        avi_monitor, skipped = convert_monitor_entity(key, f5_monitor)
        if skipped:
            add_status_row('monitor', f5_monitor["type"], key, 'partial',
                           skipped, avi_monitor)
        else:
            add_status_row('monitor', f5_monitor["type"], key, 'successful',
                           None, avi_monitor)
        monitor_list.append(avi_monitor)
    return monitor_list


def get_key_cert_obj(name, key_file_name, cert_file_name, folder_path, option):
    """
    Read key and cert files from given location and construct avi
    SSLKeyAndCertificate objects
    :param name: SSLKeyAndCertificate object name
    :param key_file_name: key file name
    :param cert_file_name: cert file name
    :param folder_path: location of key and cert files
    :param option: api-upload or cli-file both requires different
    object structure
    :return:SSLKeyAndCertificate object
    """
    key = None
    cert = None
    key_file_name = key_file_name.replace('\"', '')
    cert_file_name = cert_file_name.replace('\"', '')
    folder_path = folder_path+os.path.sep
    try:
        key_file = open(folder_path+key_file_name, "r")
        key = key_file.read()
    except:
        LOG.error("Error to read file" + folder_path+key_file_name)
    try:
        cert_file = open(folder_path+cert_file_name, "r")
        cert = cert_file.read()
    except:
        LOG.error("Error to read file" + folder_path+cert_file_name)
    ssl_kc_obj = None
    if key and cert:
        if option == "cli-upload":
            cert = {"certificate": cert}
        ssl_kc_obj = {
                'name': name,
                'key': key,
                'certificate': cert,
                'key_passphrase': ''
            }
    return ssl_kc_obj


def update_with_default_profile(profile_type, profile, profile_config):
    """
    Profiles can have inheritance used by attribute defaults-from in F5
    configuration this method recursively gets all the attributes from the
    default objects and forms complete object
    :param profile_type: type of profile
    :param profile: currant profile object
    :param profile_config: F5 profile config dict
    :return: Complete profile with updated attributes from defaults
    """
    parent_name = profile.get("defaults from", None)
    if parent_name:
        parent_profile = profile_config.get(profile_type + " " +
                                            parent_name, None)
        if parent_profile:
            parent_profile = get_defaults(profile_type, parent_profile)
            parent_profile = copy.deepcopy(parent_profile)
            parent_profile.update(profile)
            profile = parent_profile
    return profile


def convert_http_profile(profile, name):
    supported_attr = ["insert xforwarded for", "xff alternative names",
                      "max header size", "ramcache min object size",
                      "ramcache max age", "ramcache max object size",
                      "ramcache insert age header",
                      "compress keep accept encoding",
                      "compress content type include"]
    skipped = [key for key in profile.keys()
               if key not in supported_attr]
    app_profile = dict()
    sg_obj = None
    app_profile['name'] = name
    app_profile['type'] = 'APPLICATION_PROFILE_TYPE_HTTP'
    http_profile = dict()
    http_profile['x_forwarded_proto_enabled'] = profile.get(
        'insert xforwarded for', False)
    http_profile['xff_alternate_name'] = profile.get(
        'xff alternative names', None)
    header_size = profile.get('max header size', 49152)
    http_profile['client_max_header_size'] = int(header_size)/1024
    app_profile["http_profile"] = http_profile
    cache = profile.get('ramcache', 'disable')
    if not cache == 'disable':
        cache_config = dict()
        cache_config['min_object_size'] = profile.get(
            'ramcache min object size', 100)
        cache_config['query_cacheable'] = True
        cache_config['max_object_size'] = profile.get(
            'ramcache max object size', 4194304)
        age_header = profile.get('ramcache insert age header', 'disable')
        if age_header == 'enable':
            cache_config['age_header'] = True
        else:
            cache_config['age_header'] = False
        cache_config['enabled'] = True
        cache_config['default_expire'] = profile.get('ramcache max age', 600)
        http_profile["cache_config"] = cache_config
    compression = profile.get('compress', 'disable')
    if not compression == 'disable':
        compression_profile = dict()
        compression_profile["type"] = "AUTO_COMPRESSION"
        compression_profile["compression"] = True
        encoding = profile.get("compress keep accept encoding", "disable")
        if encoding == "disable":
            encoding = True
        else:
            encoding = False
        compression_profile["remove_accept_encoding_header"] = encoding
        content_type = profile.get("compress content type include", "")
        if content_type:
            content_types = content_type.keys()+content_type.values()
            sg_obj = {"name": name+"-content_type"}
            uris = []
            for content_type in content_types:
                uri = {"str": content_type}
                uris.append(uri)
            sg_obj["uris"] = uris

            compression_profile["compressible_content_ref"]\
                = name + "-content_type"
        http_profile["compression_profile"] = compression_profile
    app_profile["http_profile"] = http_profile
    return app_profile, sg_obj, skipped


def convert_profile_config(profile_config, certs_location, option):
    """
    Converts F5 profiles to equivalent Avi profiles
    :param profile_config: F5 Profile config dict
    :param certs_location: location of cert and key file location
    :param option: api-upload or cli-file both requires different
    object structure
    :return:
    """
    ssl_key_cert_list = []
    app_profile_list = []
    ssl_profile_list = []
    pki_profile_list = []
    persist_profile_list = []
    string_group = []
    hash_algorithm = []
    network_profile_list = []
    for key in profile_config.keys():
        converted_objs = []
        profile_type, name = key.split(" ")
        profile = profile_config[key]
        profile = update_with_default_profile(profile_type,
                                              profile, profile_config)
        if profile_type in ("clientssl", "serverssl"):
            supported_attr = ["cert", "key", "ciphers", "unclean shutdown",
                              "crl file", "ca file", "defaults from"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            key_cert_obj = None
            cert_file = profile.get("cert", None)
            key_file = profile.get("key", None)
            key_file = None if key_file == 'none' else key_file
            cert_file = None if cert_file == 'none' else cert_file
            if key_file and cert_file:
                key_cert_obj = get_key_cert_obj(
                    name, key_file, cert_file, certs_location, option)
            if key_cert_obj:
                ssl_key_cert_list.append(key_cert_obj)
                converted_objs.append({'key_cert': key_cert_obj})
            ciphers = profile.get('ciphers', 'DEFAULT')
            ciphers = ciphers.replace('\"', '')
            ciphers = 'AES:3DES:RC4' if ciphers in ['DEFAULT',
                                                    'NATIVE'] else ciphers
            ssl_profile = dict()
            ssl_profile['name'] = name
            ssl_profile['accepted_ciphers'] = ciphers
            close_notify = profile.get('unclean shutdown', None)
            if close_notify and close_notify == 'enabled':
                ssl_profile['send_close_notify'] = True
            else:
                ssl_profile['send_close_notify'] = False
            ssl_profile_list.append(ssl_profile)
            converted_objs.append({'ssl_profile': ssl_profile})
            options = profile.get("options", "")
            if isinstance(options, dict):
                opt = []
                for opt_key in options.keys():
                    opt.append(opt_key+' '+options[opt_key])
                options = opt
            accepted_versions = []
            if "no tlsv1" not in options:
                accepted_versions.append({"type": "SSL_VERSION_TLS1"})
            if "no tlsv1.1" not in options:
                accepted_versions.append({"type": "SSL_VERSION_TLS1_1"})
            if "no tlsv1.2" not in options:
                accepted_versions.append({"type": "SSL_VERSION_TLS1_2"})
            if accepted_versions:
                ssl_profile["accepted_versions"] = accepted_versions
            if options:
                skipped_options = [key for key in options if key not in [
                    "no tlsv1", "no tlsv1.1", "no tlsv1.2", None]]
                skipped.append({"Unsupported options": skipped_options})

            crl_file_name = profile.get('crl file', None)
            ca_file_name = profile.get('ca file', None)
            crl_file_name = None if crl_file_name == 'none' else crl_file_name
            ca_file_name = None if ca_file_name == 'none' else ca_file_name
            if not ca_file_name and crl_file_name:
                LOG.warning("Skipped PKI profile for profile %s "
                            "because of missing CA file" % name)

            elif ca_file_name:
                ca_file_name = ca_file_name.replace('\"', '')
                pki_profile = dict()
                file_path = certs_location+os.path.sep+ca_file_name
                pki_profile["name"] = name
                error = False
                try:
                    ca_file = open(file_path, "r")
                    ca = ca_file.read()
                    pki_profile["ca_certs"] = [{'certificate': ca}]
                except:
                    LOG.error("Error to read file" +
                              certs_location+os.path.sep+ca_file_name)
                    error = True
                if crl_file_name:
                    crl_file_name = crl_file_name.replace('\"', '')
                    file_path = certs_location+os.path.sep+crl_file_name
                    try:
                        crl_file = open(file_path, "r")
                        crl = crl_file.read()
                        pki_profile["crls"] = [{'body': crl}]
                    except:
                        LOG.error("Error to read file" +
                                  certs_location+os.path.sep+crl_file_name)
                        error = True
                if not error:
                    pki_profile_list.append(pki_profile)
                    converted_objs.append({'pki_profile': pki_profile})
        elif profile_type == 'http':
            app_profile, sg_obj, skipped = convert_http_profile(profile, name)
            if sg_obj:
                string_group.append(sg_obj)
                converted_objs.append({'string_group': sg_obj})
            app_profile_list.append(app_profile)
            converted_objs.append({'app_profile': app_profile})
        elif profile_type == 'dns':
            skipped = profile.keys()
            app_profile = dict()
            app_profile['name'] = name
            app_profile['type'] = 'APPLICATION_PROFILE_TYPE_DNS'
            app_profile_list.append(app_profile)
            converted_objs.append({'app_profile': app_profile})
        elif profile_type == 'fastL4':
            supported_attr = ["idle timeout", "software syncookie",
                              "defaults from"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            syn_protection = (profile.get("software syncookie", None) ==
                              'enabled')
            timeout = profile.get("idle timeout", 0)
            ntwk_profile = {
                  "profile": {
                        "tcp_fast_path_profile": {
                          "session_idle_timeout": timeout,
                          "enable_syn_protection": syn_protection
                        },
                        "type": "PROTOCOL_TYPE_TCP_FAST_PATH"
                  },
                  "name": name
                }
            network_profile_list.append(ntwk_profile)
            converted_objs.append({'network_profile': ntwk_profile})
        elif profile_type == 'fasthttp':
            supported_attr = ["idle timeout", "defaults from"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            timeout = profile.get("idle-timeout", 0)
            ntwk_profile = {
                "profile": {
                    "tcp_proxy_profile": {
                        "idle_connection_timeout": timeout
                    },
                    "type": "PROTOCOL_TYPE_TCP_PROXY"
                },
                "name": name
            }
            network_profile_list.append(ntwk_profile)
            converted_objs.append({'network_profile': ntwk_profile})
        elif profile_type == 'tcp':
            supported_attr = [" idle timeout", "defaults from"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            timeout = profile.get("idle timeout", 0)
            ntwk_profile = {
                  "profile": {
                        "tcp_proxy_profile": {
                          "session_idle_timeout": timeout
                        },
                        "type": "PROTOCOL_TYPE_TCP_PROXY"
                  },
                  "name": name
                }
            network_profile_list.append(ntwk_profile)
            converted_objs.append({'network_profile': ntwk_profile})
        elif profile_type == 'udp':
            supported_attr = ["idle timeout", "datagram lb", "defaults from"]
            skipped = [key for key in profile.keys()
                       if key not in supported_attr]
            per_pkt = profile.get("datagram lb", 'disable')
            timeout = profile.get("idle timeout", 0)
            ntwk_profile = {
                "profile": {
                    "type": "PROTOCOL_TYPE_UDP_FAST_PATH",
                    "udp_fast_path_profile": {
                        "per_pkt_loadbalance": (per_pkt == 'enable'),
                        "session_idle_timeout": timeout
                    }
                },
                "name": "System-UDP-Fast-Path"
            }
            network_profile_list.append(ntwk_profile)
            converted_objs.append({'network_profile': ntwk_profile})
        elif profile_type == 'persist':
            persist_mode = profile.get("mode")
            if persist_mode == "cookie":
                supported_attr = ["cookie name", "mode", "defaults from"]
                skipped = [key for key in profile.keys()
                           if key not in supported_attr]
                cookie_name = profile.get("cookie name", None)
                if cookie_name == "none":
                    cookie_name = None
                persist_profile = {
                    "server_hm_down_recovery": "HM_DOWN_PICK_NEW_SERVER",
                    "http_cookie_persistence_profile": {
                        "cookie_name": cookie_name
                    },
                    "persistence_type": "PERSISTENCE_TYPE_HTTP_COOKIE",
                    "name": name
                }
            elif persist_mode == "ssl":
                supported_attr = ["mode", "defaults from"]
                skipped = [key for key in profile.keys()
                           if key not in supported_attr]
                persist_profile = {
                    "server_hm_down_recovery": "HM_DOWN_PICK_NEW_SERVER",
                    "persistence_type": "PERSISTENCE_TYPE_TLS",
                    "name": name
                }
            elif persist_mode == "source addr":
                supported_attr = ["timeout", "mode", "defaults from"]
                skipped = [key for key in profile.keys()
                           if key not in supported_attr]
                timeout = profile.get("timeout", 180)
                if timeout > 0:
                    timeout = int(timeout)/60
                persist_profile = {
                    "server_hm_down_recovery": "HM_DOWN_PICK_NEW_SERVER",
                    "persistence_type": "PERSISTENCE_TYPE_CLIENT_IP_ADDRESS",
                    "ip_persistence_profile": {
                        "ip_persistent_timeout": timeout
                    },
                    "name": name
                }
            elif persist_mode == "hash":
                hash_algorithm.append(name)
                add_status_row('profile', "hash-persistence", name,
                               'indirect-mapping', None,
                               "Will be mapped to pools lb algorithm")
                continue
            else:
                LOG.error(
                    'persist mode not supported skipping conversion: ' + name)
                continue
            persist_profile_list.append(persist_profile)
            converted_objs.append({'persist_profile': persist_profile})
        else:
            LOG.warning("Not supported profile type: %s" % profile_type)
            add_status_row('profile', profile_type, name, 'skipped', None, None)
            continue
        if skipped:
            add_status_row('profile', profile_type, name, 'partial',
                           skipped, converted_objs)
        else:
            add_status_row('profile', profile_type, name, 'successful',
                           skipped, converted_objs)
    avi_profiles = dict()
    avi_profiles["ssl_key_cert_list"] = ssl_key_cert_list
    avi_profiles["app_profile_list"] = app_profile_list
    avi_profiles["ssl_profile_list"] = ssl_profile_list
    avi_profiles["pki_profile_list"] = pki_profile_list
    avi_profiles["network_profile_list"] = network_profile_list
    avi_profiles["persist_profile_list"] = persist_profile_list
    return avi_profiles, string_group, hash_algorithm


def get_profiles_for_vs(profiles, profile_config, tenant):
    """
    Searches for profile refs in converted profile config if not found creates
    default profiles
    :param profiles: profiles in f5 config assigned to VS
    :param profile_config: avi profile config
    :param tenant: tenant name for default profiles
    :return: returns list of profile refs assigned to VS in avi config
    """
    vs_ssl_profile_names = []
    pool_ssl_profile_names = []
    app_profile_names = []
    network_profile_names = []
    if not profiles:
        profiles = dict()
    if isinstance(profiles, str):
        profiles = profiles.replace(" {}", "")
        profiles = {profiles: None}
    for name in profiles.keys():
        ssl_profiles = [obj for obj in profile_config["ssl_profile_list"]
                        if obj['name'] == name]
        if ssl_profiles or name in ["clientssl", "serverssl"]:
            if not ssl_profiles and name == "clientssl":
                vs_ssl_profile_names.append({
                    "profile": tenant + ":System-Standard", "cert": None})
                continue
            if not ssl_profiles and name == "serverssl":
                pool_ssl_profile_names.append({
                    "profile": tenant+":System-Standard", "cert": None})
                continue
            key_cert = [obj for obj in profile_config["ssl_key_cert_list"]
                        if obj['name'] == name]
            key_cert = name if key_cert else None
            profile = profiles.get(name, None)
            keys = profile.keys()
            if "clientside" in keys:
                vs_ssl_profile_names.append({"profile": name, "cert": key_cert})
            elif "serverside" in keys:
                pool_ssl_profile_names.append(
                    {"profile": name, "cert": key_cert})
        app_profiles = [obj for obj in profile_config["app_profile_list"]
                        if obj['name'] == name]
        if app_profiles or name in ["http", "dns",
                                    "web-acceleration", "http-compression"]:
            if not app_profiles and name in ("http", "fasthttp"):
                app_profile_names.append(tenant+":System-HTTP")
            elif not app_profiles and name == "dns":
                app_profile_names.append(tenant+":System-DNS")
            else:
                app_profile_names.append(name)
        network_profiles = [obj for obj in profile_config[
            "network_profile_list"] if obj['name'] == name]
        if network_profiles or name in ["fasthttp", "tcp", "udp"]:
            if not network_profiles and name == "fasthttp":
                network_profile_names.append(tenant+":System-TCP-Proxy")
            if not network_profiles and name == "fastL4":
                network_profile_names.append(tenant+":System-TCP-Fast-Path")
            elif not network_profiles and name == "tcp":
                network_profile_names.append(tenant+":System-TCP-Proxy")
            elif not network_profiles and name == "udp":
                network_profile_names.append(tenant+":System-UDP-Fast-Path")
            else:
                network_profile_names.append(name)
    if not app_profile_names:
        if network_profile_names:
            app_profile_names.append(tenant+":System-L4-Application")
        else:
            app_profile_names.append(tenant+":System-HTTP")

    return vs_ssl_profile_names, pool_ssl_profile_names, app_profile_names, \
           network_profile_names


def update_service(port, vs, enable_ssl):
    """
    iterates over services of existing vs in converted list to update
    services for port overlapping scenario
    :param port: port for currant VS
    :param vs: VS from converted config list
    :param enable_ssl: value to put in service object
    :return: boolean if service is updated or not
    """
    service_updated = False
    for service in vs['services']:
        port_end = service.get('port_range_end', None)
        if port_end and (service['port'] <= int(port) <= port_end):
            if port not in [1, 65535]:
                new_end = service['port_range_end']
                service['port_range_end'] = int(port)-1
                new_service = {'port': int(port)+1,
                               'port_range_end': new_end,
                               'enable_ssl': enable_ssl}
                vs['services'].append(new_service)
            elif port == 1:
                service['port'] = 2
            elif port == 65535:
                service['port_range_end'] = 65534
            service_updated = True
            break
    return service_updated


def get_service_obj(destination, vs_list, enable_ssl):
    """
    Checks port overlapping scenario for port value 0 in F5 config
    :param destination: IP and Port destination of VS
    :param vs_list: List of existing vs converted to avi config
    :param enable_ssl: value to put in service objects
    :return: List of services for VS
    """
    parts = destination.split(':')
    ip_addr = parts[0]
    port = parts[1] if len(parts) == 2 else 80
    if port == 'any':
        port = 0
    if isinstance(port, str) and (not port.isdigit()):
        port = get_port_by_protocol(port)

    vs_dup_ips = [vs for vs in vs_list if vs['ip_address']['addr'] == ip_addr]
    if int(port) > 0:
        for vs in vs_dup_ips:
            service_updated = update_service(port, vs, enable_ssl)
            if service_updated:
                break
        services_obj = [{'port': port, 'enable_ssl': enable_ssl}]
    else:
        used_ports = []
        for vs in vs_dup_ips:
            for service in vs['services']:
                used_ports.append(service['port'])
        if used_ports:
            services_obj = []
            if 65535 not in used_ports:
                used_ports.append(65536)
            used_ports = sorted(used_ports, key=int)
            start = 1
            for i in range(len(used_ports)):
                if start == used_ports[i]:
                    start += 1
                    continue
                end = int(used_ports[i])-1
                services_obj.append({'port': start,
                                     'port_range_end': end,
                                     'enable_ssl': enable_ssl})
                start = int(used_ports[i])+1
        else:
            services_obj = [{'port': 1, 'port_range_end': 65535,
                             'enable_ssl': enable_ssl}]
    return services_obj, ip_addr


def clone_pool(pool_name, vs_name, avi_pool_list):
    """
    If pool is shared with other VS pool is cloned for other VS as Avi dose not
    support shared pools with new pool name as <pool_name>-<vs_name>
    :param pool_name: Name of the pool to be cloned
    :param vs_name: Name of the VS for pool to be cloned
    :param avi_pool_list: new pool to be added to this list
    :return: new pool object
    """
    new_pool = None
    for pool in avi_pool_list:
        if pool["name"] == pool_name:
            new_pool = copy.deepcopy(pool)
            break
    if new_pool:
        new_pool["name"] = pool_name+"-"+vs_name
        avi_pool_list.append(new_pool)
        return new_pool["name"]


def update_pool_for_persist(avi_pool_list, pool_ref, persist_profile,
                            tenant, hash_profiles, persist_config):
    pool_updated = True
    pool_obj = [pool for pool in avi_pool_list if pool["name"] == pool_ref]
    pool_obj = pool_obj[0]
    persist_profile_obj = [obj for obj in persist_config
                           if obj["name"] == persist_profile]
    persist_ref_key = "application_persistence_profile_ref"
    if persist_profile_obj:
        pool_obj[persist_ref_key] = persist_profile
    elif persist_profile == "hash" or persist_profile in hash_profiles:
        del pool_obj["lb_algorithm"]
        hash_algorithm = "LB_ALGORITHM_CONSISTENT_HASH_SOURCE_IP_ADDRESS"
        pool_obj["lb_algorithm_hash"] = hash_algorithm
    elif persist_profile == "cookie":
        pool_obj[persist_ref_key] = tenant + "System-Persistence-Http-Cookie"
    elif persist_profile == "ssl":
        pool_obj[persist_ref_key] = tenant + "System-Persistence-TLS"
    elif persist_profile == "source addr":
        pool_obj[persist_ref_key] = tenant + "System-Persistence-Client-IP"
    else:
        pool_updated = False
    return pool_updated


def add_ssl_to_pool(avi_pool_list, pool_ref, pool_ssl_profiles, tenant):
    """
    F5 serverside SSL need to be added to pool if VS contains serverside SSL
    profile this method add that profile to pool
    :param avi_pool_list: List of pools to search pool object
    :param pool_ref: name of the pool
    :param pool_ssl_profiles: ssl profiles to be added to pool
    :param tenant: tenant name for default object
    """
    if pool_ssl_profiles["profile"] == "serverssl":
        pool_ssl_profiles = tenant+":"+"System-Default-Cert"
    for pool in avi_pool_list:
        if pool_ref == pool["name"]:
            if pool_ssl_profiles["cert"]:
                pool["ssl_key_and_certificate_ref"] = pool_ssl_profiles
            if not pool.get("ssl_profile_ref", None):
                pool["ssl_profile_ref"] = "admin:System-Standard"


def convert_vs_config(vs_config, vs_state, tenant, avi_pool_list,
                      profile_config, hash_profiles):
    """
    F5 virtual server object conversion to Avi VS object
    :param vs_config: F5 virtual server config list
    :param vs_state: state of new VS to be created in Avi
    :param tenant: tenant name for default object
    :param avi_pool_list: List of pools to handle shared pool scenario
    :param profile_config: Avi profile config for profiles referenced in vs
    :param hash_profiles: Hash profiles handled separately as
    mapped to lb algorithm
    :return: List of Avi VS configs
    """
    vs_list = []
    supported_attr = ['profiles', 'destination', 'pool', 'persist']
    for vs_name in vs_config.keys():
        f5_vs = vs_config[vs_name]
        skipped = [key for key in f5_vs.keys() if key not in supported_attr]
        enabled = (vs_state == 'enable')
        vs_profiles = f5_vs.get("profiles", None)
        ssl_vs, ssl_pool, app_prof, ntwk_prof = get_profiles_for_vs(
            vs_profiles, profile_config, tenant)
        enable_ssl = False
        if ssl_vs:
            enable_ssl = True
            if app_prof[0] == (tenant+':System-HTTP'):
                app_prof[0] = tenant+':System-Secure-HTTP'
        destination = f5_vs["destination"]
        services_obj, ip_addr = get_service_obj(destination, vs_list,
                                                enable_ssl)

        pool_ref = f5_vs.get("pool", None)
        if pool_ref:
            shared_vs = [obj for obj in vs_list
                         if obj.get("pool_ref", "") == pool_ref]
            if shared_vs:
                pool_ref = clone_pool(pool_ref, vs_name, avi_pool_list)
            if ssl_pool:
                add_ssl_to_pool(avi_pool_list, pool_ref, ssl_pool[0], tenant)
            persist_ref = (f5_vs.get("persist", None))
            if persist_ref:
                persist_config = profile_config.get("persist_profile_list",
                                                    None)
                pool_updated = update_pool_for_persist(
                    avi_pool_list, pool_ref, persist_ref, tenant,
                    hash_profiles, persist_config)
                if not pool_updated:
                    skipped.append("persist")
                    LOG.warning("persist type not supported skipped for vs:" +
                                vs_name)
        vs_obj = {
            'name': vs_name,
            'type': 'VS_TYPE_NORMAL',
            'ip_address': {
                'addr': ip_addr,
                'type': 'V4'
            },
            'enabled': enabled,
            'services': services_obj,
            'application_profile_ref': app_prof[0],
            'pool_ref': pool_ref
        }
        if ntwk_prof:
            vs_obj['network_profile_ref'] = ntwk_prof[0]
        if enable_ssl:
            default_cert = tenant+":"+"System-Default-Cert"
            vs_obj['ssl_profile_name'] = ssl_vs[0]["profile"]
            if ssl_vs[0]["cert"]:
                vs_obj['ssl_key_and_certificate_refs'] = [ssl_vs[0]["cert"]]
            else:
                vs_obj['ssl_key_and_certificate_refs'] = [default_cert]

        vs_list.append(vs_obj)
        if skipped:
            add_status_row('virtual', None, vs_name, 'partial', skipped, vs_obj)
        else:
            add_status_row('virtual', None, vs_name, 'successful',
                           skipped, vs_obj)

    return vs_list


def add_status_row(f5_type, f5_sub_type, f5_id, status, skipped_params,
                   avi_object):
    global row_count
    global ws
    ws['A'+str(row_count)] = f5_type
    ws['B'+str(row_count)] = f5_sub_type
    ws['C'+str(row_count)] = f5_id
    ws['D'+str(row_count)] = status
    ws['E'+str(row_count)] = str(skipped_params)
    ws['F'+str(row_count)] = str(avi_object)
    row_count += 1


def convert_to_avi_dict(f5_config_dict, output_file_path,
                        vs_state, certs_location, tenant, option):
    global row_count
    global ws
    wb = Workbook()
    ws = wb.active
    ws['A'+str(row_count)] = 'F5 type'
    ws['B'+str(row_count)] = 'F5 SubType'
    ws['C'+str(row_count)] = 'F5 ID'
    ws['D'+str(row_count)] = 'Status'
    ws['E'+str(row_count)] = 'Skipped settings'
    ws['F'+str(row_count)] = 'Avi Object'
    row_count += 1
    avi_config_dict = {}
    monitor_config_list = convert_monitor_config(f5_config_dict.get(
        "monitor", {}))
    del f5_config_dict["monitor"]
    avi_config_dict["HealthMonitor"] = monitor_config_list
    LOG.debug("Converted health monitors")
    avi_pool_list = convert_pool_config(f5_config_dict.get("pool", {}), tenant,
                                        monitor_config_list)
    del f5_config_dict["pool"]
    avi_config_dict["Pool"] = avi_pool_list
    LOG.debug("Converted pools")
    f5_profile_dict = f5_config_dict.get("profile", {})
    avi_profiles, string_group, hash_profiles = convert_profile_config(
        f5_profile_dict, certs_location, option)
    del f5_config_dict["profile"]
    avi_config_dict["SSLKeyAndCertificate"] = avi_profiles["ssl_key_cert_list"]
    avi_config_dict["SSLProfile"] = avi_profiles["ssl_profile_list"]
    avi_config_dict["PKIProfile"] = avi_profiles["pki_profile_list"]
    avi_config_dict["ApplicationProfile"] = avi_profiles["app_profile_list"]
    avi_config_dict["NetworkProfile"] = avi_profiles["network_profile_list"]
    avi_config_dict["ApplicationPersistenceProfile"] = avi_profiles[
        "persist_profile_list"]
    avi_config_dict["StringGroup"] = string_group
    LOG.debug("Converted ssl profiles")
    avi_vs_list = convert_vs_config(f5_config_dict.get("virtual", {}), vs_state,
                                    tenant, avi_pool_list, avi_profiles,
                                    hash_profiles)
    avi_config_dict["VirtualService"] = avi_vs_list
    del f5_config_dict["virtual"]
    LOG.debug("Converted VS")
    for f5_type in f5_config_dict.keys():
        f5_obj = f5_config_dict[f5_type]
        for key in f5_obj.keys():
            sub_type = None
            if ' ' in key:
                sub_type, key = key.rsplit(' ', 1)
            add_status_row(f5_type, sub_type, key, 'skipped', None, None)
    wb.save(filename = output_file_path+os.path.sep+"ConversionStatus.xls")
    return avi_config_dict