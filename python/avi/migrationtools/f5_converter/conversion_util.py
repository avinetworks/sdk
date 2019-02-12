import copy
import logging
import os
import pandas
import ast
import re
import random
import avi.migrationtools.f5_converter.converter_constants as conv_const

from xlsxwriter import Workbook
from openpyxl import load_workbook
from pkg_resources import parse_version
from avi.migrationtools.avi_migration_utils import MigrationUtil, tenants

LOG = logging.getLogger(__name__)
csv_writer_dict_list = []

# Added variable for checking progress and get overall object.
ppcount = 0
ptotal_count = 0
global fully_migrated
fully_migrated = 0
used_pool_groups = {}
used_pool = {}

class F5Util(MigrationUtil):

    def get_conv_status(self, skipped, indirect_list, ignore_dict, f5_object,
                        user_ignore=None, na_list=None):
        """
        Update skipped list for conversion status
        :param skipped: All skipped attributes after conversion
        :param indirect_list: List of attrs to be mapped as indirect mapping
        :param ignore_dict: Dict of default values for column skipped for defaults
        :param f5_object: Currant F5 object
        :param user_ignore: List of attributes user wants not to be shown in skipped
        :param na_list: List of attributes marked as not applicable
        :return: Conversion status dict
        """
        conv_status = dict()
        user_ignore = [] if not user_ignore else user_ignore
        na_list = [] if not na_list else na_list

        conv_status['user_ignore'] = [val for val in skipped if
                                      val in user_ignore]
        skipped = [attr for attr in skipped if attr not in user_ignore]

        conv_status['indirect'] = [val for val in skipped if
                                   val in indirect_list]
        skipped = [attr for attr in skipped if attr not in indirect_list]

        conv_status['na_list'] = [val for val in skipped if val in na_list]
        skipped = [attr for attr in skipped if attr not in na_list]

        default_skip = []
        for key in ignore_dict.keys():
            f5_val = f5_object.get(key)
            default_val = ignore_dict.get(key)
            if key in skipped and f5_val == default_val:
                default_skip.append(key)
        if default_skip:
            skipped = [attr for attr in skipped if attr not in default_skip]

        conv_status['skipped'] = skipped
        conv_status['default_skip'] = default_skip
        if skipped:
            status = conv_const.STATUS_PARTIAL
        else:
            status = conv_const.STATUS_SUCCESSFUL
        conv_status['status'] = status
        return conv_status

    def get_avi_pool_down_action(self, action):
        """
        Maps Pool down action from F5 config to Avi Config
        :param action: F5 action string
        :return: Avi action String
        """
        action_close_con = {
            "type": "FAIL_ACTION_CLOSE_CONN"
        }
        if action == "reset":
            return action_close_con
        if action == "reselect":
            return action_close_con
        else:
            return action_close_con

    def get_cc_algo_val(self, cc_algo):
        """
        congestion-control algorithm conversion
        :param cc_algo: F5 algorithm value
        :return: Avi algorithm value
        """
        avi_algo_val = "CC_ALGO_NEW_RENO"
        if cc_algo == "high-speed":
            avi_algo_val = "CC_ALGO_HTCP"
        elif cc_algo == "cubic":
            avi_algo_val = "CC_ALGO_CUBIC"
        return avi_algo_val

    def add_conv_status(self, f5_type, f5_sub_type, f5_id, conv_status,
                        avi_object=None, need_review=None):
        """
        Adds as status row in conversion status csv
        :param f5_type: Object type
        :param f5_sub_type: Object sub type
        :param f5_id: Name oconv_f object
        :param conv_status: dict of conversion status
        :param avi_object: Converted objectconverted avi object
        """
        global csv_writer_dict_list
        # Added space if f5_sub_type None for pivot table
        row = {
            'F5 type': f5_type,
            'F5 SubType': f5_sub_type if f5_sub_type else ' ',
            'F5 ID': f5_id,
            'Status': conv_status.get('status', ''),
            'Skipped settings': str(conv_status.get('skipped', '')),
            'Skipped for defaults': str(conv_status.get('default_skip', '')),
            'Indirect mapping': str(conv_status.get('indirect', '')),
            'Not Applicable': str(conv_status.get('na_list', '')),
            'User Ignored': str(conv_status.get('user_ignore', '')),
            'Avi Object': str(avi_object),
            'Needs Review': need_review
        }
        csv_writer_dict_list.append(row)

    def add_status_row(self, f5_type, f5_sub_type, f5_id, status, avi_obj=None):
        """
        Adds as status row in conversion status csv
        :param f5_type: Object type
        :param f5_sub_type: Object sub type
        :param f5_id: Name of object
        :param status: conversion status
        :param avi_obj: Converted avi object
        """
        global csv_writer_dict_list
        # Added space if f5_sub_type None for pivot table
        row = {
            'F5 type': f5_type,
            'F5 SubType': f5_sub_type if f5_sub_type else ' ',
            'F5 ID': f5_id,
            'Status': status
        }
        if avi_obj:
            row.update({
                'Avi Object': str(avi_obj)
            })
        csv_writer_dict_list.append(row)

    def add_complete_conv_status(self, output_dir, avi_config, report_name,
                                 vs_level_status):

        global csv_writer_dict_list
        global ptotal_count
        for status in conv_const.STATUS_LIST:
            status_list = [row for row in csv_writer_dict_list if
                           row['Status'] == status]
            print '%s: %s' % (status, len(status_list))
        print "Writing Excel Sheet For Converted Configuration..."
        ptotal_count = ptotal_count + len(csv_writer_dict_list)
        if vs_level_status:
            self.vs_per_skipped_setting_for_references(avi_config)
            self.correct_vs_ref(avi_config)
        else:
            # Update the complexity level of VS as Basic or Advanced
            self.vs_complexity_level()
        self.write_status_report_and_pivot_table_in_xlsx(
            output_dir, report_name, vs_level_status)

    def get_port_by_protocol(self, protocol):
        """
        Instead of default ports for protocols F5 config has protocol in
        destination value for Avi object need to conver it to port number
        :param protocol: protocol name
        :return: integer value for protocol
        """
        if protocol == 'http':
            port = conv_const.HTTP_PORT
        elif protocol == "https":
            port = conv_const.HTTPS_PORT
        elif protocol == "ftp":
            port = conv_const.FTP_PORT
        elif protocol == "smtp":
            port = conv_const.SMTP_PORT
        elif protocol == "snmp":
            port = conv_const.SNMP_PORT
        elif protocol == "telnet":
            port = conv_const.TELNET_PORT
        elif protocol == "snmp-trap":
            port = conv_const.SNMP_TRAP_PORT
        elif protocol == "ssh":
            port = conv_const.SSH_PORT
        elif protocol == "xfer":
            port = conv_const.XFER_PORT
        elif protocol == "pcsync-https":
            port = conv_const.PCSYNC_HTTPS_PORT
        elif protocol == "macromedia-fcs":
            port = conv_const.MACROMEDIA_FCS_PORT
        elif protocol == 'imap':
            port = conv_const.IMAP_PORT
        elif protocol == 'pop3':
            port = conv_const.POP3_PORT
        elif protocol == "any":
            port = None
        else:
            return None
        return port

    def update_skip_duplicates(self, obj, obj_list, obj_type, converted_objs,
                               name, default_profile_name, merge_object_mapping,
                               ent_type, prefix, syslist):

        """
        Merge duplicate profiles
        :param obj: Source object to find duplicates for
        :param obj_list: List of object to search duplicates in
        :param obj_type: Type of object to add in converted_objs status
        :param converted_objs: Converted avi object or merged object name
        :param name: Name of the object
        :param default_profile_name : Name of root parent default profile
        :param merge_object_mapping: merged object mappings
        :param ent_type: Entity type
        :param prefix: object name prefix
        :param syslist: System object list
        :return:
        """
        dup_of = None
        if isinstance(merge_object_mapping, dict):
            merge_object_mapping[obj_type].update({name: name})
        # root default profiles are skipped for merging
        if not name == default_profile_name or obj_type == 'ssl_profile':
            dup_of, old_name = \
                self.check_for_duplicates(obj, obj_list, obj_type,
                                          merge_object_mapping, ent_type,
                                          prefix, syslist)
        if dup_of:
            converted_objs.append({obj_type: "Duplicate of %s" % dup_of})
            LOG.info(
                "Duplicate profiles: %s merged in %s" % (obj['name'], dup_of))
            if isinstance(merge_object_mapping, dict):
                if old_name in merge_object_mapping[obj_type].keys():
                    merge_object_mapping[obj_type].update({old_name: dup_of})
                merge_object_mapping[obj_type].update({name: dup_of})
        else:
            obj_list.append(obj)
            converted_objs.append({obj_type: obj})

    def get_content_string_group(self, name, content_types, tenant):
        """
        Creates Avi String group object
        :param name: name of string group
        :param content_types: list of content type
        :param tenant: tenant name to add tenant reference
        :return:
        """
        sg_obj = {"name": name + "-content_type", "type": "SG_TYPE_STRING"}
        kv = []
        for content_type in content_types:
            if content_type is None:
                LOG.warning('%s content_types %s has none', name, content_types)
                continue
            uri = {"key": content_type}
            kv.append(uri)
        sg_obj["kv"] = kv
        # Changed tenant ref to new reference.
        sg_obj['tenant_ref'] = self.get_object_ref(tenant, 'tenant')
        return sg_obj

    def get_vs_ssl_profiles(self, profiles, avi_config, prefix,
                            merge_object_mapping, sys_dict, f5_config):
        """
        Searches for profile refs in converted profile config if not found
        creates default profiles
        :param profiles: profiles in f5 config assigned to VS
        :param avi_config: converted avi config
        :param prefix: prefix for objects
        :param merge_object_mapping: Merged object mappings
        :param sys_dict: System object dict
        :return: returns list of profile refs assigned to VS in avi config
        """
        # f5_profiles = f5_config.get("profile", {})
        vs_ssl_profile_names = []
        pool_ssl_profile_names = []
        if not profiles:
            return vs_ssl_profile_names, pool_ssl_profile_names
        if isinstance(profiles, str):
            profiles = profiles.replace(" {}", "")
            profiles = {profiles: None}
        for key in profiles.keys():
            # Called tenant ref to get object name.
            tenant, name = self.get_tenant_ref(key)
            if prefix:
                name = prefix + '-' + name
            ssl_profile_list = avi_config.get("SSLProfile", [])
            sys_ssl = sys_dict['SSLProfile']
            ssl_profiles = [ob for ob in sys_ssl if ob['name'] ==
                            merge_object_mapping['ssl_profile'].get(name)
                            ] or [obj for obj in ssl_profile_list
                                  if (obj['name'] == name or name in
                                      obj.get("dup_of", []))]
            if ssl_profiles:
                cert_name = ssl_profiles[0].get('cert_name', None)
                if not cert_name:
                    cert_name = name
                ssl_key_cert_list = avi_config.get("SSLKeyAndCertificate", [])
                sys_key_cert = sys_dict['SSLKeyAndCertificate']
                key_cert = [ob for ob in sys_key_cert if ob['name'] ==
                            merge_object_mapping['ssl_cert_key'].get(cert_name)
                            ] or [obj for obj in ssl_key_cert_list if
                                  (obj['name'] == cert_name or obj['name'] ==
                                   cert_name + '-dummy' or cert_name in
                                   obj.get("dup_of", []))]
                # key_cert = key_cert[0]['name'] if key_cert else None
                if key_cert:
                    key_cert = self.get_object_ref(
                        key_cert[0]['name'], 'sslkeyandcertificate',
                        tenant=self.get_name(key_cert[0]['tenant_ref']))
                profile = profiles[key]
                context = profile.get("context") if profile else None
                if (not context) and isinstance(profile, dict):
                    if 'serverside' in profile:
                        context = 'serverside'
                    elif 'clientside' in profile:
                        context = 'clientside'
                pki_list = avi_config.get("PKIProfile", [])
                syspki = sys_dict['PKIProfile']
                pki_profiles = [ob for ob in syspki if ob['name'] ==
                                merge_object_mapping['pki_profile'].get(
                                    name)] or \
                               [obj for obj in pki_list if
                                (obj['name'] == name or
                                 name in obj.get("dup_of", []))]
                pki_profile = pki_profiles[0]['name'] if pki_profiles else None
                mode = 'SSL_CLIENT_CERTIFICATE_NONE'
                if pki_profile:
                    mode = pki_profiles[0].pop('mode',
                                               'SSL_CLIENT_CERTIFICATE_NONE')
                    pki_profile = self.get_object_ref(
                        pki_profiles[0]["name"], 'pkiprofile',
                        tenant=self.get_name(pki_profiles[0]['tenant_ref']))
                if context == "clientside":
                    ssl_prof_ref = self.get_object_ref(
                        ssl_profiles[0]["name"], 'sslprofile',
                        tenant=self.get_name(ssl_profiles[0]['tenant_ref']))
                    vs_ssl_profile_names.append({"profile": ssl_prof_ref,
                                                 "cert": key_cert,
                                                 "pki": pki_profile,
                                                 'mode': mode})
                elif context == "serverside":
                    ssl_prof_ref = self.get_object_ref(
                        ssl_profiles[0]["name"], 'sslprofile',
                        tenant=self.get_name(ssl_profiles[0]['tenant_ref']))
                    pool_ssl_profile_names.append(
                        {"profile": ssl_prof_ref, "cert": key_cert,
                         "pki": pki_profile, 'mode': mode})
        return vs_ssl_profile_names, pool_ssl_profile_names

    def get_vs_app_profiles(self, profiles, avi_config, tenant_ref, prefix,
                            oc_prof, enable_ssl, merge_object_mapping,
                            sys_dict):
        """
        Searches for profile refs in converted profile config if not found
        creates default profiles
        :param profiles: profiles in f5 config assigned to VS
        :param avi_config: converted avi config
        :param tenant_ref: Tenant referance
        :param prefix: prefix for objects
        :param oc_prof: one connect profile
        :param enable_ssl: VS ssl enabled flag
        :param merge_object_mapping: Merged object mappings
        :param sys_dict: System object dict

        :return: returns list of profile refs assigned to VS in avi config
        """
        app_profile_refs = []
        app_prof_conf = dict()
        app_profile_list = avi_config.get("ApplicationProfile", [])
        unsupported_profiles = avi_config.get('UnsupportedProfiles', [])
        sys_app = sys_dict['ApplicationProfile']
        if not profiles:
            profiles = {}
        if isinstance(profiles, str):
            profiles = profiles.replace(" {}", "")
            profiles = {profiles: None}
        for name in profiles.keys():
            # Called tenant ref to get object name.
            name = self.get_tenant_ref(name)[1]
            # Added prefix for objects
            if prefix:
                name = '%s-%s' % (prefix, name)
            app_profiles = [ob for ob in sys_app if ob['name'] ==
                            merge_object_mapping['app_profile'].get(name)] or [
                               obj for obj in app_profile_list if
                               (obj['name'] == name
                                or name in obj.get("dup_of", []))]
            if app_profiles:
                app_prof_name = app_profiles[0]['name']
                app_profile_refs.append(self.get_object_ref(
                    app_prof_name, 'applicationprofile',
                    tenant=self.get_name(app_profiles[0]['tenant_ref'])))

                if app_profiles[0].get('HTTPPolicySet', None):
                    app_prof_conf['policy_name'] = app_profiles[0]['HTTPPolicySet']
                if app_profiles[0].get('fallback_host', None):
                    app_prof_conf['f_host'] = app_profiles[0]['fallback_host']
                # prerequisite user need to create default auth profile
                if app_profiles[0].get('realm', None):
                    app_prof_conf['realm'] = {
                        "type": "HTTP_BASIC_AUTH",
                        "auth_profile_ref": self.get_object_ref(
                            'System-Default-Auth-Profile', 'authprofile',
                            tenant=self.get_name(
                                app_profiles[0]['tenant_ref'])),
                        "realm": app_profiles[0]['realm']
                    }

        if not app_profile_refs:
            not_supported = [key for key in profiles.keys() if
                             key in unsupported_profiles]
            if not_supported:
                LOG.warning(
                    'Profiles not supported by Avi : %s' % not_supported)
                return app_prof_conf
            if enable_ssl:
                app_profile_refs.append(
                    self.get_object_ref('System-SSL-Application',
                                        'applicationprofile', tenant='admin'))
                app_prof_conf['app_prof'] = app_profile_refs
                return app_prof_conf
            else:
                app_profile_refs.append(
                    self.get_object_ref('System-L4-Application',
                                        'applicationprofile', tenant='admin'))
                app_prof_conf['app_prof'] = app_profile_refs
                return app_prof_conf
            # Added prefix for objects
            if prefix:
                value = '%s-%s' % (prefix, value)
            default_app_profile = [ob for ob in sys_app if ob['name'] ==
                                   merge_object_mapping['app_profile'].get(
                                       value)] or [
                                      obj for obj in app_profile_list if
                                      (obj['name'] == value
                                       or value in obj.get("dup_of", []))]
            tenant = self.get_name(default_app_profile[0]['tenant_ref']) if \
                default_app_profile else '/api/tenant/?name=admin'
            app_profile_refs.append(
                self.get_object_ref(default_app_profile[0]['name'],
                                    'applicationprofile', tenant=tenant))
        app_prof_conf['app_prof'] = app_profile_refs
        return app_prof_conf

    def get_vs_ntwk_profiles(self, profiles, avi_config, prefix,
                             merge_object_mapping, sys_dict):
        """
        Searches for profile refs in converted profile config if not found
        creates default profiles
        :param profiles: profiles in f5 config assigned to VS
        :param avi_config: converted avi config
        :param prefix: prefix for objects
        :param merge_object_mapping: merged object mappings
        :param sys_dict: System object dict
        :return: returns list of profile refs assigned to VS in avi config
        """
        network_profile_names = []
        if not profiles:
            return []
        if isinstance(profiles, str):
            profiles = profiles.replace(" {}", "")
            profiles = {profiles: None}
        for name in profiles.keys():
            # Called tenant method to get object name
            tenant, name = self.get_tenant_ref(name)
            # Added prefix for objects
            if prefix:
                name = prefix + '-' + name
            ntwk_prof_lst = avi_config.get("NetworkProfile")
            sysnw = sys_dict['NetworkProfile']
            network_profiles = [ob for ob in sysnw if
                                ob['name'] == merge_object_mapping[
                                    'network_profile'].get(name)] or \
                               [obj for obj in ntwk_prof_lst if (
                                       obj['name'] == name or name in
                                       obj.get("dup_of", []))]
            if network_profiles:
                network_profile_ref = self.get_object_ref(
                    network_profiles[0]['name'], 'networkprofile',
                    tenant=self.get_name(network_profiles[0]['tenant_ref']))
                network_profile_names.append(network_profile_ref)
        return network_profile_names

    def update_service(self, port, vs, enable_ssl):
        """
        iterates over services of existing vs in converted list to update
        services for port overlapping scenario
        :param port: port for currant VS
        :param vs: VS from converted config list
        :param enable_ssl: value to put in service object
        :return: boolean if service is updated or not
        """
        service_updated = False
        vs_new_service = []
        for service in vs['services']:
            port_end = service.get('port_range_end', None)
            if not port_end and int(service['port']) == int(port):
                return 'duplicate_ip_port'
            if port_end and (service['port'] <= int(port) <= port_end):
                if port not in [conv_const.PORT_START, conv_const.PORT_END]:
                    if service['port'] == int(port) == port_end:
                        return 'duplicate_ip_port'
                    elif service['port'] == int(port):
                        service['port'] = int(port) + 1
                    elif service['port_range_end'] == int(port):
                        service['port_range_end'] = int(port) - 1
                    else:
                        new_port = int(port) + 1
                        new_end = service['port_range_end']
                        service['port_range_end'] = int(port) - 1
                        new_service = {'port': new_port,
                                       'port_range_end': new_end,
                                       'enable_ssl': enable_ssl}
                        vs_new_service.append(new_service)
                elif port == conv_const.PORT_START:
                    service['port'] = 2
                elif port == conv_const.PORT_END:
                    service['port_range_end'] = (conv_const.PORT_START - 1)
                service_updated = True
                break
        vs['services'].extend(vs_new_service)
        return service_updated

    def get_service_obj(self, destination, avi_config, enable_ssl,
                        controller_version, tenant_name, cloud_name, prefix,
                        vs_name, input_vrf=None):
        """
        Checks port overlapping scenario for port value 0 in F5 config
        :param destination: IP and Port destination of VS
        :param avi_config: Dict of avi config
        :param enable_ssl: value to put in service objects
        :param controller_version: Version of controller
        :param tenant_name: Name of tenant
        :param cloud_name: Name of cloud
        :param prefix: name prefix
        :param vs_name: Name of VS
        :param input_vrf: Vrf context name
        :return: services_obj, ip_addr of vs and ref of vsvip
        """

        parts = destination.split(':')
        ip_addr = parts[0]
        ip_addr = ip_addr.strip()
        vrf = None
        # Removed unwanted string from ip address
        if '%' in ip_addr:
            ip_addr, vrf = ip_addr.split('%')
        # Added support to skip virtualservice with ip address any
        if ip_addr == 'any':
            LOG.debug("Skipped:VS with IP address: %s" % str(destination))
            return None, None, None, None
        # Added check for IP V4
        matches = re.findall('^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', ip_addr)
        if not matches or ip_addr == '0.0.0.0':
            LOG.warning(
                'Avi does not support IPv6 Generated random ipv4 for vs:'
                ' %s' % ip_addr)
            ip_addr = ".".join(map(str, (
                random.randint(0, 255) for _ in range(4))))
        port = parts[1] if len(parts) == 2 else conv_const.DEFAULT_PORT
        # Get the list of vs which shared the same vip
        if parse_version(controller_version) >= parse_version('17.1'):
            vs_dup_ips = \
                [vs for vs in avi_config['VirtualService'] if
                 vs['vip'][0]['ip_address']['addr'] ==
                 ip_addr]
        else:
            vs_dup_ips = \
                [vs for vs in avi_config['VirtualService'] if
                 vs['ip_address']['addr'] == ip_addr]

        if port == 'any':
            port = '0'
        if isinstance(port, str) and (not port.isdigit()):
            port = self.get_port_by_protocol(port)
        # Port is None then skip vs
        if not port:
            LOG.debug("Skipped:Port not supported %s" % str(parts[1]))
            return None, None, None, None
        if int(port) > 0:
            for vs in vs_dup_ips:
                service_updated = self.update_service(port, vs, enable_ssl)
                if service_updated == 'duplicate_ip_port':
                    LOG.debug('Skipped: Duplicate IP-Port for vs %s', vs_name)
                    return None, None, None, None
                if service_updated:
                    break
            services_obj = [{'port': port, 'enable_ssl': enable_ssl}]
        else:
            if {service.get('port_range_end') for vs in vs_dup_ips for
               service in vs['services']}:
                LOG.debug('Skipped: Duplicate IP-Port for vs %s', vs_name)
                return None, None, None, None
            used_ports = list({service['port'] for vs in vs_dup_ips for
                              service in vs['services']})
            if used_ports:
                services_obj = []
                if conv_const.PORT_END not in used_ports:
                    used_ports.append(conv_const.PORT_END + 1)
                used_ports = sorted(used_ports, key=int)
                start = conv_const.PORT_START
                for i in range(len(used_ports)):
                    if start == used_ports[i]:
                        start += 1
                        continue
                    end = int(used_ports[i]) - 1
                    if end < start:
                        start += 1
                        continue
                    services_obj.append({'port': start,
                                         'port_range_end': end,
                                         'enable_ssl': enable_ssl})
                    start = int(used_ports[i]) + 1
            else:
                services_obj = [
                    {'port': 1, 'port_range_end': conv_const.PORT_END,
                     'enable_ssl': enable_ssl}]
        # Getting vrf ref
        if vrf:
            self.add_vrf(avi_config, vrf, cloud_name)

        vrf_config = avi_config['VrfContext']
        vrf_ref = self.get_vrf_context_ref(destination, vrf_config,
                                           'virtual service', vs_name,
                                           cloud_name)
        if input_vrf:
            vrf_ref = self.get_object_ref(input_vrf, 'vrfcontext',
                                          cloud_name=cloud_name)

        updated_vsvip_ref = None
        if parse_version(controller_version) >= parse_version('17.1'):
            vs_vip_name = self.create_update_vsvip(
                ip_addr, avi_config['VsVip'],
                self.get_object_ref(tenant_name, 'tenant'),
                self.get_object_ref(cloud_name, 'cloud', tenant=tenant_name),
                prefix,
                vrf_ref)
            if vs_vip_name == '':
                updated_vsvip_ref = ''
            else:
                updated_vsvip_ref = self.get_object_ref(vs_vip_name, 'vsvip',
                                                        tenant_name, cloud_name)
        return services_obj, ip_addr, updated_vsvip_ref, vrf_ref

    def clone_pool(self, pool_name, clone_for, avi_pool_list, is_vs,
                   tenant=None):
        """
        If pool is shared with other VS pool is cloned for other VS as Avi dose
        not support shared pools with new pool name as <pool_name>-<vs_name>
        :param pool_name: Name of the pool to be cloned
        :param clone_for: Name of the VS for pool to be cloned
        :param avi_pool_list: new pool to be added to this list
        :param is_vs: True if this cloning is for VS
        :param tenant: if pool is shared across partition then coned for tenant
        :return: new pool object
        """
        LOG.debug("Cloning pool %s for %s " % (pool_name, clone_for))
        new_pool = None
        for pool in avi_pool_list:
            if pool["name"] == pool_name:
                new_pool = copy.deepcopy(pool)
                break
        if new_pool:
            if pool_name in used_pool:
                used_pool[pool_name] += 1
            else:
                used_pool[pool_name] = 1
            LOG.debug('Cloning Pool for %s', clone_for)
            new_pool["name"] = '{}-{}'.format(pool_name, used_pool[pool_name])
            if tenant:
                new_pool["tenant_ref"] = self.get_object_ref(tenant, 'tenant')
            if is_vs:
                # removing config added from VS config to pool
                new_pool["application_persistence_profile_ref"] = None
                new_pool["ssl_profile_ref"] = None
                new_pool["ssl_key_and_certificate_ref"] = None
                new_pool["pki_profile_ref"] = None
            avi_pool_list.append(new_pool)
            pool_ref = new_pool["name"]
            LOG.debug("Cloned pool successfully %s for %s " % (
                new_pool["name"], clone_for))
            return pool_ref

    def remove_https_mon_from_pool(self, avi_config, pool_ref, tenant, sysdict):
        pool = [p for p in avi_config['Pool'] if p['name'] == pool_ref]
        if pool:
            hm_refs = pool[0].get('health_monitor_refs', [])
            for hm_ref in hm_refs:
                hm = [h for h in (sysdict['HealthMonitor'] + avi_config[
                    'HealthMonitor']) if
                      self.get_object_ref(
                          h['name'], 'healthmonitor', tenant=tenant) == hm_ref]
                if hm and hm[0]['type'] == 'HEALTH_MONITOR_HTTPS':
                    pool[0]['health_monitor_refs'].remove(hm_ref)
                    LOG.warning(
                        'Skipping %s this reference from %s pool '
                        'because of health monitor type is HTTPS and VS '
                        'has no ssl profile.'
                        % (hm_ref, pool_ref))

    def remove_http_mon_from_pool(self, avi_config, pool_ref, tenant, sysdict):
        pool = [p for p in avi_config['Pool'] if p['name'] == pool_ref]
        if pool:
            hm_refs = pool[0].get('health_monitor_refs', [])
            for hm_ref in hm_refs:
                hm = [h for h in (sysdict['HealthMonitor'] + avi_config[
                    'HealthMonitor']) if
                      self.get_object_ref(
                          h['name'], 'healthmonitor', tenant=tenant) == hm_ref]

                if hm and hm[0]['type'] == 'HEALTH_MONITOR_HTTP':
                    pool[0]['health_monitor_refs'].remove(hm_ref)
                    LOG.warning(
                        'Skipping %s this reference from %s pool because of'
                        ' health monitor type is HTTPS and VS has no ssl '
                        'profile.' % (hm_ref, pool_ref))

    def remove_https_mon_from_pool_group(self, avi_config, poolgroup_ref,
                                         tenant, sysdict):
        poolgroup = [p for p in avi_config['PoolGroup'] if self.get_object_ref(
            p['name'], 'poolgroup', tenant=tenant) == poolgroup_ref]
        if poolgroup:
            pool_members = [p['pool_ref'] for p in poolgroup[0]['members']]
            for pool_ref in pool_members:
                pool_ref = self.get_name(pool_ref)
                self.remove_https_mon_from_pool(avi_config, pool_ref, tenant,
                                                sysdict)

    def remove_http_mon_from_pool_group(self, avi_config, poolgroup_ref, tenant,
                                        sysdict):
        poolgroup = [p for p in avi_config['PoolGroup'] if self.get_object_ref(
            p['name'], 'poolgroup', tenant=tenant) == poolgroup_ref]
        if poolgroup:
            pool_members = [p['pool_ref'] for p in poolgroup[0]['members']]
            for pool_ref in pool_members:
                pool_name = self.get_name(pool_ref)
                self.remove_http_mon_from_pool(
                    avi_config, pool_name, tenant, sysdict)

    def add_ssl_to_pool(self, avi_pool_list, pool_ref, pool_ssl_profiles,
                        tenant='admin'):
        """
        F5 serverside SSL need to be added to pool if VS contains serverside SSL
        profile this method add that profile to pool
        :param avi_pool_list: List of pools to search pool object
        :param pool_ref: name of the pool
        :param pool_ssl_profiles: ssl profiles to be added to pool
        :param tenant: tenant name
        """
        for pool in avi_pool_list:
            if pool_ref == pool["name"]:
                if pool_ssl_profiles["profile"]:
                    pool["ssl_profile_ref"] = pool_ssl_profiles["profile"]
                if pool_ssl_profiles["pki"]:
                    pool["pki_profile_ref"] = pool_ssl_profiles["pki"]
                if pool_ssl_profiles["cert"]:
                    pool["ssl_key_and_certificate_ref"] = pool_ssl_profiles[
                        "cert"]

    def add_ssl_to_pool_group(self, avi_config, pool_group_ref, ssl_pool,
                              tenant_ref):
        """

        :param avi_config:
        :param pool_group_ref:
        :param ssl_pool:
        :param tenant_ref:
        :return:
        """
        pool_group = [obj for obj in avi_config['PoolGroup']
                      if obj['name'] == pool_group_ref]
        if pool_group:
            pool_group = pool_group[0]
            for member in pool_group['members']:
                pool_name = self.get_name(member['pool_ref'])
                self.add_ssl_to_pool(
                    avi_config['Pool'], pool_name, ssl_pool, tenant_ref)

    def update_pool_for_persist(self, avi_pool_list, pool_ref, persist_profile,
                                hash_profiles, persist_config, tenant,
                                merge_object_mapping, syspersist,
                                app_prof_type):
        """
        Updates pool for persistence profile assigned in F5 VS config
        :param avi_pool_list: List of all converted pool objects to avi config
        :param pool_ref: pool name to be updated
        :param persist_profile: persistence profile to be added to pool
        :param hash_profiles: list of profile name for which pool's lb algorithm
        updated to hash
        :param persist_config: list of all converted persistence profiles
        :param tenant: tenant of a pool
        :param app_prof_type: type of application profile for the VS
        :param merge_object_mapping: merged object mapping
        :param syspersist: system persistence profile
        :return: Boolean of is pool updated successfully
        """
        pool_updated = True
        persist_type = None
        pool_obj = [pool for pool in avi_pool_list if pool["name"] == pool_ref]
        if not pool_obj:
            LOG.error("Pool %s not found to add profile %s" %
                      (pool_ref, persist_profile))
            return False, None
        pool_obj = pool_obj[0]
        persist_profile_obj = \
            [ob for ob in syspersist if ob['name'] ==
             merge_object_mapping['app_per_profile'].get(persist_profile)] or \
            [obj for obj in persist_config if (
                    obj["name"] == persist_profile or persist_profile
                    in obj.get("dup_of", []))]
        persist_ref_key = "application_persistence_profile_ref"
        if persist_profile_obj:
            if app_prof_type == 'APPLICATION_PROFILE_TYPE_L4' and (
                    persist_profile_obj[0]['persistence_type'] !=
                    'PERSISTENCE_TYPE_CLIENT_IP_ADDRESS'):
                pool_obj[persist_ref_key] = self.get_object_ref(
                                                'System-Persistence-Client-IP',
                                                'applicationpersistenceprofile')
                persist_type = 'PERSISTENCE_TYPE_CLIENT_IP_ADDRESS'
                LOG.debug("Defaulted to Client IP persistence profile for '%s' "
                          "Pool in VS of L4 app type " % pool_ref)
            else:
                obj_tenant = persist_profile_obj[0]['tenant_ref']
                pool_obj[persist_ref_key] = \
                    self.get_object_ref(
                        persist_profile_obj[0]['name'],
                        'applicationpersistenceprofile',
                        tenant=self.get_name(obj_tenant))
                persist_type = persist_profile_obj[0]['persistence_type']
        elif persist_profile == "hash" or persist_profile in hash_profiles:
            del pool_obj["lb_algorithm"]
            hash_algorithm = "LB_ALGORITHM_CONSISTENT_HASH_SOURCE_IP_ADDRESS"
            pool_obj["lb_algorithm_hash"] = hash_algorithm
        else:
            pool_updated = False
        return pool_updated, persist_type

    def update_pool_group_for_persist(
            self, avi_config, pool_ref, persist_profile, hash_profiles,
            persist_config, tenant, merge_object_mapping, syspersist,
            app_prof_type):
        pool_group_updated = True
        persist_type = None
        pool_group = [obj for obj in avi_config['PoolGroup']
                      if obj['name'] == pool_ref]
        if pool_group:
            pool_group = pool_group[0]
            for member in pool_group['members']:
                pool_name = self.get_name(member['pool_ref'])
                pool_updated, persist_type = self.update_pool_for_persist(
                    avi_config['Pool'], pool_name, persist_profile,
                    hash_profiles, persist_config, tenant, merge_object_mapping,
                    syspersist, app_prof_type)
                if not pool_updated:
                    pool_group_updated = False
        return pool_group_updated, persist_type

    def update_pool_for_fallback(self, host, avi_pool_list, pool_ref):
        """
        Update pool for fallback host config
        :param host: Redirect url
        :param avi_pool_list: List of all converted pools
        :param pool_ref: Name of the pool for which config is to be added
        """
        pool_obj = [pool for pool in avi_pool_list if pool["name"] == pool_ref]
        if pool_obj:
            pool_obj = pool_obj[0]
            fail_action = {
                "redirect":
                    {
                        "status_code": "HTTP_REDIRECT_STATUS_CODE_302",
                        "host": host,
                        "protocol": "HTTPS"
                    },
                "type": "FAIL_ACTION_HTTP_REDIRECT"
            }
            pool_obj["fail_action"] = fail_action

    def get_snat_list_for_vs(self, snat_pool):
        """
        Converts the f5 snat pool config object to Avi snat list
        :param snat_pool: f5 snat pool config
        :return: Avi snat list
        """
        snat_list = []
        members = snat_pool.get("members")
        ips = []
        if isinstance(members, dict):
            ips = members.keys() + members.values()
        elif isinstance(members, str):
            ips = [members]
        ips = [ip for ip in ips if ip]
        for ip in ips:
            # Removed unwanted string from ip address
            if '/' in ip or '%' in ip:
                ip = ip.split('/')[-1]
                ip = ip.split('%')[-2]
            snat_obj = {
                "type": "V4",
                "addr": ip
            }
            snat_list.append(snat_obj)
        return snat_list

    def cleanup_config(self, avi_config):
        self.remove_dup_key(avi_config["SSLKeyAndCertificate"])
        self.remove_dup_key(avi_config["ApplicationProfile"])
        self.remove_dup_key(avi_config["NetworkProfile"])
        self.remove_dup_key(avi_config["SSLProfile"])
        self.remove_dup_key(avi_config["PKIProfile"])
        self.remove_dup_key(avi_config["ApplicationPersistenceProfile"])
        self.remove_dup_key(avi_config["HealthMonitor"])
        self.remove_dup_key(avi_config["IpAddrGroup"])
        avi_config.pop('hash_algorithm', [])
        avi_config.pop('OneConnect', [])
        avi_config.pop('UnsupportedProfiles', [])
        for profile in avi_config['ApplicationProfile']:
            profile.pop('HTTPPolicySet', None)
            profile.pop('realm', [])
            profile.pop('fallback_host', [])
        for profile in avi_config.get('PKIProfile', []):
            profile.pop('mode', None)
        for profile in avi_config.get('SSLProfile', []):
            profile.pop('cert_name', None)
        if 'Tenant' in avi_config:
            for tenant in avi_config['Tenant']:
                if tenant['name'] == 'admin':
                    avi_config['Tenant'].remove(tenant)

    def create_hdr_erase_rule(self, name, hdr_name, rule_index):
        return self.create_header_rule(
            name, hdr_name, "HTTP_REMOVE_HDR", None, rule_index)

    def create_hdr_insert_rule(self, name, hdr_name, val, rule_index):
        return self.create_header_rule(
            name, hdr_name, "HTTP_ADD_HDR", val.strip(), rule_index)

    def create_header_rule(self, name, hdr_name, action, val,
                           rule_index):
        rule = {
            "name": name,
            "index": rule_index,
            "hdr_action": [
                {
                    "action": action,
                    "hdr": {
                        "name": hdr_name.strip(),
                        "value": {
                            "val": val
                        }
                    }
                }
            ]
        }
        return rule

    def create_network_security_rule(self, name, ip, mask, tenant):
        if '%' in ip:
            ip = ip.split('%')[0]
        rule = {
            "name": name,
            "tenant_ref": self.get_object_ref(tenant, 'tenant'),
            "rules": [
                {
                    "index": 1,
                    "enable": True,
                    "name": "Rule 1",
                    "age": 0,
                    "action": "NETWORK_SECURITY_POLICY_ACTION_TYPE_DENY",
                    "match": {
                        "client_ip": {
                            "prefixes": [
                                {
                                    "ip_addr": {
                                        "type": "V4",
                                        "addr": ip
                                    },
                                    "mask": mask
                                }
                            ],
                            "match_criteria": "IS_NOT_IN"
                        }
                    },
                    "log": False
                }
            ]
        }
        return rule

    def add_vrf(self,  avi_config, vrf, cloud_ref):
        vrf_name = 'vrf-%s' % vrf
        vrf_list = avi_config['VrfContext']
        vrf_obj = [obj for obj in vrf_list if obj['name'] == vrf_name]
        if not vrf_obj:
            vrf_obj = {
                "name": vrf_name,
                "system_default": False,
                "cloud_ref": self.get_object_ref(cloud_ref, 'cloud'),
                "tenant_ref": self.get_object_ref('admin', 'tenant')
            }
            if vrf_name == 'global':
                vrf_obj['system_default'] = True
            vrf_list.append(vrf_obj)

    def get_tenant_ref(self, name):
        tenant = 'admin'
        if name and name.startswith('/'):
            parts = name.split('/', 2)
            tenant = parts[1]
            if not parts[2]:
                LOG.warning('Invalid tenant ref : %s' % name)
            name = parts[2]
        elif name and '/' in name:
            parts = name.split('/')
            # Changed the index to get the tenant and name in case of
            # prefixed name
            tenant = parts[-2]
            name = parts[-1]
        if tenant.lower() == 'common':
            tenant = 'admin'
        if '/' in name:
            name = name.split('/')[1]
        if ' ' in tenant:
            tenant = tenant.split(' ')[-1]
        return tenant, name

    def get_app_profile_type(self, profile_name, avi_config):
        profiles = avi_config.get('ApplicationProfile', [])
        # Called tenant method to get object name
        profile_name = self.get_tenant_ref(profile_name)[1]
        profile = [obj for obj in profiles if obj['name'] == profile_name]
        if profile:
            return profile[0]['type']
        else:
            return 'APPLICATION_PROFILE_TYPE_HTTP'

    def update_pool_for_service_port(self, pool_list, pool_name, hm_list,
                                     sys_hm_list):
        rem_hm = []
        pool = [obj for obj in pool_list if obj['name'] == pool_name]
        if pool:
            pool[0]['use_service_port'] = True
            # Checking monitor ports if use_service_port is true
            if pool[0].get('health_monitor_refs'):
                for hm in pool[0]['health_monitor_refs']:
                    hm_name = self.get_name(hm)
                    hm_ob = [ob for ob in (hm_list + sys_hm_list) if
                             ob['name'] == hm_name]
                    if hm_ob and (not hm_ob[0].get('monitor_port')):
                        rem_hm.append(hm)
                        LOG.debug("Removing monitor reference of %s from pool"
                                  " %s as 'use_service_port' is true but "
                                  "monitor has no port", hm_name,
                                  pool_name)
                if rem_hm:
                    pool[0]['health_monitor_refs'] = [
                        h_monitor for h_monitor in pool[0]
                        ['health_monitor_refs'] if h_monitor not in rem_hm]

                    rem_hm = [self.get_name(hmonitor) for hmonitor in rem_hm]
                    csv_row = [cl for cl in csv_writer_dict_list if cl[
                               'F5 type'] == 'pool' and self.get_tenant_ref(
                        cl['F5 ID'])[1] == pool_name]
                    if csv_row:
                        if csv_row[0]['Skipped settings'] in ('[]', ''):
                            csv_row[0]['Skipped settings'] = str([{
                                                            'monitor': rem_hm}])
                        else:
                            init_val = eval(csv_row[0]['Skipped settings'])
                            if not isinstance(init_val, list):
                                init_val = [init_val]
                            mon_val = [
                                val['monitor'].extend(rem_hm) for val in
                                init_val if isinstance(val, dict) and
                                'monitor' in val]
                            if bool(mon_val):
                                csv_row[0]['Skipped settings'] = str(init_val)
                            else:
                                init_val.append({'monitor': rem_hm})
                                csv_row[0]['Skipped settings'] = str(init_val)
                        csv_row[0]['Status'] = conv_const.STATUS_PARTIAL
                        csv_row[0]['Avi Object'] = str({'pools': pool})

    def rreplace(self, s, old, new, occurrence):
        li = s.rsplit(old, occurrence)
        return new.join(li)

    def get_project_path(self):
        return os.path.abspath(os.path.dirname(__file__))

    def clone_pool_if_shared(self, ref, avi_config, vs_name, tenant, p_tenant,
                             persist_type, controller_version, app_prof_ref,
                             sysdict, cloud_name='Default-Cloud', prefix=None):
        """
        clones pool or pool group if its shard between multiple VS or partitions
        in F5
        :param ref: reference of pool or pool group
        :param avi_config: Avi configuration cloned pool or pool groups to be
        added
        :param vs_name: Name of the vs to be added
        :param tenant: tenant name of vs
        :param p_tenant: tenant name of pool
        :param persist_type: persistence profile type
        :param controller_version:
        :param app_prof_ref: Application profile referance
        :param sysdict:
        :param cloud_name:
        :param prefix:
        :return:
        """
        is_pool_group = False
        pool_group_obj = None
        # Added prefix for objects
        if prefix:
            ref = prefix + '-' + ref
        # Search the pool or pool group with name in avi config for the same
        # tenant as VS
        pool_obj = [pool for pool in avi_config['Pool'] if pool['name'] == ref
                    and pool['tenant_ref'] == self.get_object_ref(tenant,
                    'tenant')]
        pool_per_ref = pool_obj[0].get(
            'application_persistence_profile_ref') if pool_obj else None
        pool_per_name = self.get_name(pool_per_ref) if pool_per_ref else None
        pool_per_types = [obj['persistence_type'] for obj in (avi_config[
                          'ApplicationPersistenceProfile'] + sysdict[
                          'ApplicationPersistenceProfile']) if obj['name'] ==
                          pool_per_name] if pool_per_name else []
        pool_per_type = pool_per_types[0] if pool_per_types else None
        if not pool_obj:
            pool_group_obj = [pool for pool in avi_config['PoolGroup']
                              if pool['name'] == ref and
                              pool['tenant_ref'] == self.get_object_ref(
                    tenant, 'tenant')]
        if pool_group_obj:
            is_pool_group = True
        if p_tenant:
            shared_vs = [obj for obj in avi_config['VirtualService']
                         if obj.get("pool_ref", "") == self.get_object_ref(
                    ref, 'pool', tenant=p_tenant, cloud_name=cloud_name)]
            if not shared_vs:
                shared_vs = [obj for obj in avi_config['VirtualService']
                             if obj.get("pool_group_ref", "") ==
                             self.get_object_ref(
                                 ref, 'poolgroup', tenant=p_tenant,
                                 cloud_name=cloud_name)]
        else:
            shared_vs = [obj for obj in avi_config['VirtualService']
                         if obj.get("pool_ref", "") == self.get_object_ref(
                    ref, 'pool', tenant=tenant, cloud_name=cloud_name)]
            if not shared_vs:
                shared_vs = [obj for obj in avi_config['VirtualService']
                             if obj.get("pool_group_ref", "") ==
                             self.get_object_ref(
                                 ref, 'poolgroup', tenant=tenant,
                                 cloud_name=cloud_name)]
        if not tenant == p_tenant:
            if is_pool_group:
                ref = self.clone_pool_group(ref, vs_name, avi_config, True,
                                            tenant, cloud_name=cloud_name)
            else:
                ref = self.clone_pool(ref, vs_name, avi_config['Pool'],
                                      True, tenant)
        if shared_vs:
            if is_pool_group:
                ref = self.clone_pool_group(ref, vs_name, avi_config, True,
                                            tenant, cloud_name=cloud_name)
            else:
                shared_appref = shared_vs[0].get('application_profile_ref')
                shared_apptype = None
                if shared_appref:
                    shared_appname = self.get_name(shared_appref)
                    shared_appobjs = [ob for ob in (avi_config[
                                      'ApplicationProfile'] + sysdict[
                                      'ApplicationProfile']) if ob['name'] ==
                                      shared_appname]
                    shared_appobj = shared_appobjs[0] if shared_appobjs else {}
                    shared_apptype = shared_appobj['type'] if shared_appobj \
                        else None
                app_prof_name = self.get_name(app_prof_ref)
                app_prof_objs = [appob for appob in (avi_config[
                                 'ApplicationProfile'] + sysdict[
                                 'ApplicationProfile']) if appob['name'] ==
                                 app_prof_name]
                app_prof_obj = app_prof_objs[0] if app_prof_objs else {}
                app_prof_type = app_prof_obj['type'] if app_prof_obj else None

                if self.is_pool_clone_criteria(
                        controller_version, app_prof_type, shared_apptype,
                        persist_type, pool_per_type, shared_appobj,
                        app_prof_obj):
                    LOG.debug('Cloned the pool %s for VS %s', ref, vs_name)
                    ref = self.clone_pool(ref, vs_name, avi_config['Pool'],
                                          True, tenant)
                else:
                    LOG.debug("Shared pool %s for VS %s", ref, vs_name)

        return ref, is_pool_group

    def is_pool_clone_criteria(self, controller_version, app_prof_type,
                               shared_apptype, persist_type, pool_per_type,
                               shared_appobj, app_prof_obj):
        if parse_version(controller_version) < parse_version(
           '17.1.6') or app_prof_type != 'APPLICATION_PROFILE_TYPE_HTTP' \
           or shared_apptype != app_prof_type or (
                persist_type and persist_type !=
                'PERSISTENCE_TYPE_HTTP_COOKIE') or (
                pool_per_type and pool_per_type !=
                'PERSISTENCE_TYPE_HTTP_COOKIE') or (
                shared_appobj.get('http_profile', {}).get(
                    'connection_multiplexing_enabled') != app_prof_obj.get(
                    'http_profile', {}).get('connection_multiplexing_enabled') or (
                shared_appobj.get('http_profile', {}).get(
                    'cache_config') != app_prof_obj.get(
                    'http_profile', {}).get('cache_config'))):
            return True
        else:
            return False

    def clone_pool_group(self, pool_group_name, clone_for, avi_config, is_vs,
                         tenant='admin', cloud_name='Default-Cloud'):
        """
        If pool is shared with other VS pool is cloned for other VS as Avi dose
        not support shared pools with new pool name as <pool_name>-<vs_name>
        :param pool_group_name: Name of the pool group to be cloned
        :param clone_for: Name of the object/entity for pool group to be cloned
        :param avi_config: new pool to be added to avi config
        :param is_vs: True if clone is called for VS
        :param tenant: if f5 pool is shared across partition then coned for
        tenant
        :param cloud_name:
        :return: new pool group name
        """
        pg_ref = None
        new_pool_group = None
        for pool_group in avi_config['PoolGroup']:
            if pool_group["name"] == pool_group_name:
                new_pool_group = copy.deepcopy(pool_group)
                break
        if new_pool_group:
            if pool_group_name in used_pool_groups:
                used_pool_groups[pool_group_name] += 1
            else:
                used_pool_groups[pool_group_name] = 1
            LOG.debug('Cloning pool group for %s', clone_for)
            new_pool_group["name"] = '{}-{}'.format(
                pool_group_name, used_pool_groups[pool_group_name])
            pg_ref = new_pool_group["name"]
            new_pool_group["tenant_ref"] = self.get_object_ref(tenant, 'tenant')
            avi_config['PoolGroup'].append(new_pool_group)
            for member in new_pool_group['members']:
                pool_name = self.get_name(member['pool_ref'])
                pool_name = self.clone_pool(pool_name, clone_for,
                                            avi_config['Pool'], is_vs, tenant)
                member['pool_ref'] = self.get_object_ref(
                    pool_name, 'pool', tenant=tenant, cloud_name=cloud_name)
        return pg_ref

    def add_tenants(self, avi_config_dict):
        if tenants:
            avi_config_dict['Tenant'] = []
            for tenant in tenants:
                avi_config_dict['Tenant'].append({
                    'name': tenant,
                    'local': True
                })

    def get_cell_format(self, workbook, cell_format_info):
        format_col = cell_format_info['col']
        format = workbook.add_format(cell_format_info['fromat'])
        return format_col, format

    def write_status_report_and_pivot_table_in_xlsx(
            self, output_dir, report_name, vs_level_status):
        """
        This function defines that add status sheet and pivot table sheet in
        xlsx format
        :param output_dir: Path of output directory
        :param report_name: filename to write report
        :param vs_level_status: Flag to include VS wise detailed status or not
        :return: None
        """
        global ppcount
        global ptotal_count
        # List of fieldnames for headers
        if vs_level_status:
            fieldnames = ['F5 type', 'F5 SubType', 'F5 ID', 'Status',
                          'Skipped settings', 'Indirect mapping',
                          'Not Applicable', 'User Ignored',
                          'Skipped for defaults', 'Complexity Level',
                          'VS Reference', 'Overall skipped settings',
                          'Avi Object', 'Needs Review']
        else:
            fieldnames = ['F5 type', 'F5 SubType', 'F5 ID', 'Status',
                          'Skipped settings', 'Indirect mapping',
                          'Not Applicable',
                          'User Ignored', 'Skipped for defaults',
                          'Complexity Level', 'Avi Object', 'Needs Review']

        # xlsx workbook
        report_path = output_dir + os.path.sep + "%s-ConversionStatus.xlsx" % \
                                                 report_name
        status_wb = Workbook(report_path)
        # xlsx worksheet
        status_ws = status_wb.add_worksheet("Status Sheet")
        # Lock the first row of xls report.
        status_ws.freeze_panes(1, 0)
        first_row = 0
        for header in fieldnames:
            col = fieldnames.index(header)
            status_ws.write(first_row, col, header)
        row = 1
        for row_data in csv_writer_dict_list:
            ppcount += 1
            for _key, _value in row_data.items():
                col = fieldnames.index(_key)
                status_ws.write(row, col, _value)
            # Added call for progress function.
            msg = "excel sheet conversion started..."
            self.print_progress_bar(ppcount, ptotal_count, msg,
                                    prefix='Progress', suffix='')
            row += 1
        status_wb.close()
        # create dataframe for row list
        df = pandas.DataFrame(csv_writer_dict_list, columns=fieldnames)
        # create pivot table using pandas
        pivot_table = \
            pandas.pivot_table(df, index=["Status", "F5 type", "F5 SubType"],
                               values=[], aggfunc=[len], fill_value=0)
        # create dataframe for pivot table using pandas
        pivot_df = pandas.DataFrame(pivot_table)
        master_book = \
            load_workbook(report_path)
        master_writer = pandas.ExcelWriter(report_path, engine='openpyxl')
        master_writer.book = master_book
        # Add pivot table in Pivot sheet
        pivot_df.to_excel(master_writer, 'Pivot Sheet')
        master_writer.save()

    def format_string_to_json(self, avi_string):
        """
        This function defines that it convert string into json format to
        convert into dict
        :param avi_string: string to be converted
        :return: Return converted string
        """
        avi_string = avi_string.split('__/__')[0]
        return ast.literal_eval(avi_string)

    def get_csv_object_list(self, csv_writer_dict_list, command_list):
        """
        This method is used for getting csv object
        :param csv_writer_dict_list: CSV row of object from xlsx report
        :param command_list: List of netscaler commands
        :return: List of CSV rows
        """

        csv_object = [row for row in csv_writer_dict_list if
                      row['Status'] in [conv_const.STATUS_PARTIAL,
                                        conv_const.STATUS_SUCCESSFUL] and
                      '->' not in row['Avi Object'] and
                      row['F5 type'] in command_list]
        return csv_object

    def get_and_update_csv_row(self, csv_object, vs_ref):
        """
        This function defines that update csv row.
        :param csv_object: csv object
        :param vs_ref: Name of VS
        :return: Skipped attribute list
        """

        if 'VS Reference' in csv_object and \
                vs_ref not in csv_object['VS Reference']:
            csv_object['VS Reference'] += ',' + vs_ref
        else:
            csv_object['VS Reference'] = vs_ref
        repls = ('[', ''), (']', '')
        skipped_setting_csv = reduce(
            lambda a, kv: a.replace(*kv), repls, csv_object['Skipped settings'])
        if skipped_setting_csv:
            return [skipped_setting_csv]

    def get_csv_skipped_list(self, csv_objects, name_of_object, vs_ref,
                             field_key=None):
        """
        This method is used for getting skipped list from vs.
        :param csv_objects: CSV row of object from xlsx report
        :param name_of_object: Name of object
        :param vs_ref: Name of VS
        :param field_key: Key fromm avi json which is specific for object type
        :return: Return skipped attribute list
        """

        for csv_object in csv_objects:
            avi_objects = self.format_string_to_json(csv_object['Avi Object'])
            if isinstance(avi_objects, dict):
                avi_objects = [avi_objects]
            if not avi_objects:
                avi_objects = []
            for avi_object_json in avi_objects:
                object_found = False
                if field_key:
                    if field_key in avi_object_json and 'Duplicate' not in \
                            avi_object_json[field_key] and (
                            avi_object_json[field_key]['name'] ==
                            name_of_object):
                        object_found = True
                else:
                    if avi_object_json.get('name') and \
                                    avi_object_json['name'] == name_of_object:
                        object_found = True

                if object_found:
                    return self.get_and_update_csv_row(csv_object, vs_ref)

    def get_ssl_profile_skipped(self, profile_csv_list, ssl_profile_ref,
                                vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param profile_csv_list: List of profile(F5 type) csv rows
        :param ssl_profile_ref: Reference of ssl profile
        :param vs_ref: Name of VS
        :return: ssl profile name and skipped sttribute list
        """

        ssl_profile_name = self.get_name(ssl_profile_ref)
        skipped_list = self.get_csv_skipped_list(
            profile_csv_list, ssl_profile_name, vs_ref, field_key='ssl_profile')
        return ssl_profile_name, skipped_list

    def get_application_profile_skipped(self, profile_csv_list, app_profile_ref,
                                        vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param profile_csv_list: List of profile(F5 type) csv rows
        :param app_profile_ref: Reference of application profile
        :param vs_ref: Name of VS
        :return: application profile name and skipped sttribute list
        """

        app_profile_name = self.get_name(app_profile_ref)
        skipped_list = self.get_csv_skipped_list(
            profile_csv_list, app_profile_name, vs_ref, field_key='app_profile')
        return app_profile_name, skipped_list

    def get_network_profile_skipped(self, profile_csv_list, network_profile_ref,
                                    vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param profile_csv_list: List of profile(F5 type) csv rows
        :param network_profile_ref: Reference of Network profile
        :param vs_ref: Name of VS
        :return: network profile name and skipped sttribute list
        """

        network_profile_name = self.get_name(network_profile_ref)
        skipped_list = self.get_csv_skipped_list(
            profile_csv_list, network_profile_name, vs_ref,
            field_key='network_profile')
        return network_profile_name, skipped_list

    def get_policy_set_skipped(self, profile_csv_list, policy_set_ref, vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param profile_csv_list: List of profile(F5 type) csv rows
        :param policy_set_ref: Reference of policy set
        :param vs_ref: Name of VS
        :return: policy set name and skipped sttribute list
        """

        policy_set_name = self.get_name(policy_set_ref)
        skipped_list = self.get_csv_skipped_list(
            profile_csv_list, policy_set_name, vs_ref, field_key='policy_set')
        return policy_set_name, skipped_list

    def get_app_persistence_profile_skipped(self, csv_writer_dict_list,
                                            pool_object, vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param csv_writer_dict_list: List of csv rows
        :param pool_object: object of pool
        :param vs_ref: Name of VS
        :return: profile name and skipped attribute list
        """

        app_persistence_profile_name = self.get_name(
            pool_object['application_persistence_profile_ref'])
        csv_object = self.get_csv_object_list(csv_writer_dict_list,
                                              ['persistence'])
        skipped_list = self.get_csv_skipped_list(
            csv_object, app_persistence_profile_name, vs_ref,
            field_key='app_per_profile')
        return app_persistence_profile_name, skipped_list

    def get_pool_skipped(self, csv_objects, pool_name, vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param csv_objects: CSV row of object from xlsx report
        :param pool_name: Name of pool
        :param vs_ref: Name of VS
        :return: Skipped list of csv row
        """

        for csv_object in csv_objects:
            avi_object = self.format_string_to_json(csv_object['Avi Object'])
            if 'pools' in avi_object:
                pool_object = [pool for pool in avi_object['pools']
                               if pool['name'] == pool_name]
                if pool_object:
                    return self.get_and_update_csv_row(csv_object, vs_ref)

    def get_pool_skipped_list(self, avi_config, pool_group_name, csv_pool_rows,
                              csv_writer_dict_list, vs_ref, profile_csv_list):
        """
        This method is used for getting pool skipped list.
        :param avi_config: AVI dict
        :param pool_group_name: Name of Pool group
        :param csv_pool_rows: List of pool(F5 type) csv rows
        :param csv_writer_dict_list: List of F5 csv rows
        :param vs_ref: Name of VS
        :param profile_csv_list: List of profile(F5 type) csv rows
        :return:
        """

        pool_group_objects = [pool_group_object for pool_group_object in
                              avi_config['PoolGroup'] if
                              pool_group_object['name']
                              == pool_group_name]
        pool_members = pool_group_objects[0]['members']
        skipped_setting = {
            'pools': []
        }
        for pool_member in pool_members:
            pool_name = self.get_name(pool_member['pool_ref'])
            self.get_skipped_pool(
                avi_config, pool_name, csv_pool_rows, csv_writer_dict_list,
                vs_ref, profile_csv_list, skipped_setting)
        if skipped_setting['pools']:
            return skipped_setting

    def vs_complexity_level(self):
        """
        This method calculate the complexity of vs.
        :return:
        """
        # Get the VS object list which is having status successful and partial.
        vs_csv_objects = [row for row in csv_writer_dict_list
                          if row['Status'] in [conv_const.STATUS_PARTIAL,
                                               conv_const.STATUS_SUCCESSFUL]
                          and row['F5 type'] == 'virtual']
        for vs_csv_object in vs_csv_objects:
            virtual_service = self.format_string_to_json(
                vs_csv_object['Avi Object'])
            # Update the complexity level of VS as Basic or Advanced
            self.update_vs_complexity_level(vs_csv_object, virtual_service)

    def vs_per_skipped_setting_for_references(self, avi_config):
        """
        This functions defines that Add the skipped setting per VS CSV row
        :param avi_config: this method use avi_config for checking vs skipped
        :return: None
        """

        # Get the count of vs fully migrated
        global fully_migrated
        global ptotal_count
        global ppcount
        fully_migrated = 0
        # Get the VS object list which is having status successful and partial.
        vs_csv_objects = [row for row in csv_writer_dict_list
                          if row['Status'] in [conv_const.STATUS_PARTIAL,
                                               conv_const.STATUS_SUCCESSFUL]
                          and row['F5 type'] == 'virtual']
        # Get the list of csv rows which has profile as F5 Type
        profile_csv_list = self.get_csv_object_list(
            csv_writer_dict_list, ['profile'])
        ptotal_count = ptotal_count + len(vs_csv_objects)
        for vs_csv_object in vs_csv_objects:
            ppcount += 1
            skipped_setting = {}
            virtual_service = self.format_string_to_json(
                vs_csv_object['Avi Object'])
            # Update the complexity level of VS as Basic or Advanced
            self.update_vs_complexity_level(vs_csv_object, virtual_service)
            vs_ref = virtual_service['name']
            repls = ('[', ''), (']', '')
            # Get list of skipped setting attributes
            skipped_setting_csv = reduce(lambda a, kv: a.replace(*kv), repls,
                                         vs_csv_object['Skipped settings'])
            if skipped_setting_csv:
                skipped_setting['virtual_service'] = [skipped_setting_csv]
            # Get the skipped list for ssl key and cert
            if 'ssl_key_and_certificate_refs' in virtual_service:
                for ssl_key_and_certificate_ref in \
                        virtual_service['ssl_key_and_certificate_refs']:
                    ssl_key_cert = self.get_name(ssl_key_and_certificate_ref)
                    ssl_kc_skip = self.get_csv_skipped_list(
                        profile_csv_list, ssl_key_cert, vs_ref,
                        field_key='ssl_cert_key')
                    if ssl_kc_skip:
                        skipped_setting['ssl cert key'] = {}
                        skipped_setting['ssl cert key']['name'] = ssl_key_cert
                        skipped_setting['ssl cert key'][
                            'skipped_list'] = ssl_kc_skip

            # Get the skipped list for ssl profile name.
            # Changed ssl profile name to ssl profile ref.
            if 'ssl_profile_ref' in virtual_service:
                name, skipped = self.get_ssl_profile_skipped(
                    profile_csv_list, virtual_service['ssl_profile_ref'],
                    vs_ref)
                if skipped:
                    skipped_setting['ssl profile'] = {}
                    skipped_setting['ssl profile']['name'] = name
                    skipped_setting['ssl profile']['skipped_list'] = skipped
            # Get the skipped list for pool group.
            if 'pool_group_ref' in virtual_service:
                pool_group_name = self.get_name(
                    virtual_service['pool_group_ref'])
                csv_pool_rows = self.get_csv_object_list(csv_writer_dict_list,
                                                         ['pool'])
                pool_group_skipped_settings = self.get_pool_skipped_list(
                    avi_config, pool_group_name, csv_pool_rows,
                    csv_writer_dict_list, vs_ref, profile_csv_list)
                if pool_group_skipped_settings:
                    skipped_setting['Pool Group'] = pool_group_skipped_settings
            # Get the skipped list for pool.
            if 'pool_ref' in virtual_service:
                pool_skipped_settings = {'pools': []}
                pool_name = self.get_name(virtual_service['pool_ref'])
                csv_pool_rows = self.get_csv_object_list(csv_writer_dict_list,
                                                         ['pool'])
                self.get_skipped_pool(
                    avi_config, pool_name, csv_pool_rows, csv_writer_dict_list,
                    vs_ref, profile_csv_list, pool_skipped_settings)
                if pool_skipped_settings['pools']:
                    skipped_setting['Pool'] = pool_skipped_settings
            # Get the skipepd list for http policy.
            if 'http_policies' in virtual_service:
                policy_csv_list = self.get_csv_object_list(
                    csv_writer_dict_list, ['policy', 'profile'])
                for http_ref in virtual_service['http_policies']:
                    policy_set_name, skipped_list = self.get_policy_set_skipped(
                        policy_csv_list, http_ref['http_policy_set_ref'],
                        vs_ref)
                    if skipped_list:
                        skipped_setting['Httppolicy'] = {}
                        skipped_setting['Httppolicy']['name'] = policy_set_name
                        skipped_setting['Httppolicy'][
                            'skipped_list'] = skipped_list
                    # Get the http policy name
                    pool_csv_rows = \
                        self.get_csv_object_list(csv_writer_dict_list, ['pool'])
                    for each_http_policy in avi_config['HTTPPolicySet']:
                        if each_http_policy['name'] == policy_set_name:
                            for http_req in each_http_policy[
                              'http_request_policy']['rules']:
                                if http_req.get('switching_action', {}):
                                    self.get_skip_pools_policy(
                                        policy_set_name, http_req,
                                        avi_config, pool_csv_rows, vs_ref,
                                        profile_csv_list, skipped_setting)

            # # Get the skipped list for application_profile_ref.
            if 'application_profile_ref' in virtual_service and 'admin:System' \
                    not in virtual_service['application_profile_ref']:
                name, skipped = self.get_application_profile_skipped(
                    profile_csv_list,
                    virtual_service['application_profile_ref'],
                    vs_ref)
                if skipped:
                    skipped_setting['Application profile'] = {}
                    skipped_setting['Application profile'][
                        'name'] = name
                    skipped_setting['Application profile'][
                        'skipped_list'] = skipped
            # # Get the skipped list for network profile ref.
            if 'network_profile_ref' in virtual_service and 'admin:System' \
                    not in virtual_service['network_profile_ref']:
                name, skipped = self.get_network_profile_skipped(
                    profile_csv_list, virtual_service['network_profile_ref'],
                    vs_ref)
                if skipped:
                    skipped_setting['Network profile'] = {}
                    skipped_setting['Network profile'][
                        'name'] = name
                    skipped_setting['Network profile'][
                        'skipped_list'] = skipped
            # Update overall skipped setting of VS csv row
            if skipped_setting:
                vs_csv_object.update(
                    {'Overall skipped settings': str(skipped_setting)})
            else:
                vs_csv_object.update(
                    {'Overall skipped settings': "FULLY MIGRATION"})
                fully_migrated += 1
            # Added call for progress function.
            msg = "excel sheet conversion started..."
            self.print_progress_bar(ppcount, ptotal_count, msg,
                                    prefix='Progress', suffix='')
        csv_objects = [row for row in csv_writer_dict_list
                       if row['Status'] in [
                           conv_const.STATUS_PARTIAL,
                           conv_const.STATUS_SUCCESSFUL]
                       and row['F5 type'] != 'virtual']

        # Update the vs reference not in used if objects are not attached to
        # VS directly or indirectly
        for row in csv_objects:
            if 'VS Reference' not in row or row['VS Reference'] == '':
                row['VS Reference'] = conv_const.STATUS_NOT_IN_USE

    def create_update_vsvip(self, vip, vsvip_config, tenant_ref, cloud_ref,
                            prefix, vrf_ref):
        """
        This functions defines that create or update VSVIP object.
        :param vip: vip of VS
        :param vsvip_config: List of vs object
        :param tenant_ref: tenant reference
        :param cloud_ref: cloud reference
        :param prefix: Name prefix
        :param vrf_ref: VRF reference
        :return: None
        """

        name = vip + '-vsvip'
        # Added prefix for objects
        if prefix:
            name = '%s-%s' % (prefix, name)
        # Get the exsting vsvip object list if present
        vsvip = [vip_obj for vip_obj in vsvip_config if vip_obj['name'] == name
                 and vip_obj.get('vrf_context_ref') == vrf_ref]
        if vsvip:
            diff_ten = [vips for vips in vsvip if vips['tenant_ref'] !=
                        tenant_ref]
            if diff_ten:
                LOG.debug('VsVip %s is repeated with vrf %s but different '
                          'tenant %s', name, self.get_name(vrf_ref) if vrf_ref
                          else 'None', self.get_name(tenant_ref))
                name = ''
        # If VSVIP object not present then create new VSVIP object.
        else:
            vsvip_object = {
                "name": name,
                "tenant_ref": tenant_ref,
                "cloud_ref": cloud_ref,
                "vip": [
                    {
                        "vip_id": "0",
                        "ip_address": {
                            "type": "V4",
                            "addr": vip
                        }
                    }
                ],
            }
            if vrf_ref:
                vsvip_object["vrf_context_ref"] = vrf_ref
            vsvip_config.append(vsvip_object)

        return name

    def update_static_route(self, route):
        """
        This function defines that convert convert static routes
        :param route: Object of net static route
        :return: Return static route object
        """
        msg = None
        next_hop_ip = route.get('gw', route.get('gateway'))
        if next_hop_ip and '%' in next_hop_ip:
            next_hop_ip = next_hop_ip.split('%')[0]

        ip_addr = route.get('network', None)
        vrf = None
        # Get the mask from subnet mask
        if ip_addr and '%' in ip_addr:
            ip_addr, vrf = ip_addr.split('%')
            vrf = 'vrf-' + (
                    '/' in vrf and vrf.split('/')[0] or vrf) if vrf else None
        if ip_addr and '/' in ip_addr:
            ip_addr = ip_addr.split('/')[0]

        # set subnet mask to 0.0.0.0 if its equal to default
        if not ip_addr or ip_addr == 'default':
            ip_addr = '0.0.0.0'

        mask = sum([bin(int(x)).count('1') for x in ip_addr.split('.')])
        if next_hop_ip and ip_addr:
            static_route = {
                "route_id": 1,
                "prefix": {
                    "ip_addr": {
                        "type": "V4",
                        "addr": ip_addr
                    },
                    "mask": mask
                },
                "next_hop": {
                    "type": "V4",
                    "addr": next_hop_ip
                }
            }
            return static_route, vrf, msg
        else:
            msg = "Next hop ip is not present" if not next_hop_ip else (
                "Ip Address is not present")
            LOG.debug(msg)
            return None, None, msg

    def get_vrf_context_ref(self, f5_entity_mem, vrf_config, entity_string,
                            entity_name, cloud):
        """
        Searches for vrf context refs in converted pool config
        :param f5_entity_mem: f5 entity or object like pool
        :param vrf_config: converted vrf config
        :param entity_string: entity string
        :param entity_name: name of f5 entity
        :param cloud: name of the cloud
        :return: returns list of vrf refs assigned to entity in avi config
        """
        vrf_ref = None
        f5_entity_mem = ':' in f5_entity_mem and f5_entity_mem.split(':')[0] \
                        or f5_entity_mem if f5_entity_mem else None
        vrf = 'vrf-' + f5_entity_mem.split('%')[1] \
            if f5_entity_mem and '%' in f5_entity_mem else None
        vrf_obj = [obj for obj in vrf_config if vrf and obj["name"] == vrf]
        if vrf_obj:
            vrf_ref = self.get_object_ref(
                vrf_obj[0]['name'], 'vrfcontext', cloud_name=cloud)
        else:
            LOG.warning("VRF not found for %s %s" % (entity_string,
                                                     entity_name))
        return vrf_ref

    def net_to_static_route(self, f5_config, avi_config):
        """
        This method converts the net route to static routes and updates the
        VrfContext objects
        :param f5_config: parsed f5 config
        :param avi_config: converted config in avi
        :return:
        """
        net_config = f5_config.get('route', {})
        avi_vrf = avi_config["VrfContext"]
        # Convert net static route to vrf static route
        for key, route in net_config.iteritems():
            LOG.debug("Starting conversion from net route to static for '%s'"
                      % key)
            static_route, vrf, msg = self.update_static_route(route)
            if static_route:
                for obj in avi_vrf:
                    if obj['name'] == vrf or (not vrf and obj['name'] ==
                       'global'):
                        if obj.get('static_routes'):
                            rid = max(
                                [i['route_id'] for i in obj['static_routes']])
                            static_route['route_id'] = rid + 1
                            obj['static_routes'].append(static_route)
                        else:
                            obj['static_routes'] = [static_route]
                LOG.debug("Conversion completed for route '%s'" % key)
                self.add_conv_status(
                    'route', None, key,
                    {'status': conv_const.STATUS_SUCCESSFUL},
                    [{'route': static_route}]
                )
            else:
                LOG.debug("Conversion unsuccessful for route '%s'" % key)
                self.add_conv_status('route', None, key,
                                     {'status': conv_const.STATUS_SKIPPED}, msg)

    def update_monitor_ssl_ref(self, avi_dict, merge_obj_dict, sysdict):
        """
        This method updates the first ssl profile reference from merge
        perspective in monitors, which get attached at the time of creation
        :param avi_dict: avi configuration dict
        :param merge_obj_dict: dict having merge objects
        :param sysdict: system object dicts
        :return:
        """
        for obj in avi_dict['HealthMonitor']:
            obj_ref = obj.get('https_monitor', {}).get('ssl_attributes',
                                                       {}).get(
                'ssl_profile_ref')
            if obj_ref:
                name = self.get_name(obj_ref)
                if name in merge_obj_dict['ssl_profile']:
                    updated_name = merge_obj_dict['ssl_profile'][name]
                    prof = [ob for ob in (sysdict['SSLProfile'] + avi_dict[
                        'SSLProfile']) if ob['name'] == updated_name]
                    tenant = self.get_name(prof[0]['tenant_ref'])
                    type_cons = conv_const.OBJECT_TYPE_SSL_PROFILE
                    obj['https_monitor']['ssl_attributes']['ssl_profile_ref'] =\
                        self.get_object_ref(updated_name, type_cons, tenant)

    def update_app_profile(self, aviconfig, sys_dict):
        """
        This method updates the application profile to http when there are
        multiple services to a L4 app VS in which one of them is ssl enabled
        :param aviconfig: avi config dict
        :param sys_dict: system config dict
        :return:
        """
        for vs_obj in aviconfig['VirtualService']:
            if vs_obj.get('services') and len(vs_obj['services']) > 1 and \
                    vs_obj.get('application_profile_ref'):
                app_profile = self.get_name(vs_obj['application_profile_ref'])
                app_profile_obj = [app for app in sys_dict[
                        'ApplicationProfile'] + aviconfig['ApplicationProfile']
                                   if app['name'] == app_profile]
                if app_profile_obj and app_profile_obj[0]['type'] == \
                        'APPLICATION_PROFILE_TYPE_L4':
                    for service in vs_obj['services']:
                        if service['enable_ssl']:
                            vs_obj['application_profile_ref'] = \
                                self.get_object_ref(
                                'System-HTTP',
                                conv_const.OBJECT_TYPE_APPLICATION_PROFILE)
                            LOG.debug('Changed the application profile '
                                      'reference from L4 to System-HTTP')
                            if vs_obj.get('network_profile_ref'):
                                self.update_nw_profile(
                                    vs_obj, sys_dict, aviconfig)
                            break

    def update_nw_profile(self, vs_obj, sys_dict, aviconfig):
        nw_profile = self.get_name(vs_obj['network_profile_ref'])
        nw_profile_obj = [nw for nw in sys_dict['NetworkProfile'] +
                          aviconfig['NetworkProfile'] if nw['name'] ==
                          nw_profile]
        if nw_profile_obj and nw_profile_obj[0]['profile']['type'] != \
                'PROTOCOL_TYPE_TCP_PROXY':
            LOG.debug(
                'Changed the network profile reference from %s to '
                'TCP-Proxy for VS %s' % (nw_profile_obj[0]['profile']['type'],
                                         vs_obj['name']))
            vs_obj['network_profile_ref'] = self.get_object_ref(
                'System-TCP-Proxy', conv_const.OBJECT_TYPE_NETWORK_PROFILE)

    def set_pool_group_vrf(self, pool_ref, vrf_ref, avi_config):
        """
        This method will set vrf_ref for all pools in poolgroup
        :param pool_ref: pool group name
        :param vrf_ref: vrf ref of VS
        :param avi_config: avi config json
        :return:
        """
        pg_obj = [poolgrp for poolgrp in avi_config['PoolGroup'] if
                  poolgrp['name'] == pool_ref]
        if pg_obj:
            for member in pg_obj[0]['members']:
                poolname = self.get_name(member.get('pool_ref'))
                self.set_pool_vrf(poolname, vrf_ref, avi_config)

    def set_pool_vrf(self, pool_ref, vrf_ref, avi_config):
        """
        This method will set vrf_ref for pool
        :param pool_ref: pool name
        :param vrf_ref: vrf ref of VS
        :param avi_config: avi config json
        :return:
        """
        pool_obj = [pool for pool in avi_config['Pool'] if pool['name'] ==
                    pool_ref]
        if pool_obj and not pool_obj[0].get('vrf_ref'):
            pool_obj[0]['vrf_ref'] = vrf_ref
            LOG.debug("Added vrf ref to the pool %s", pool_ref)

    def clone_http_policy_set(self, policy, vs_name, avi_config, tenant_name,
                              cloud_name):
        """
        This function clones policy which is shared with more than one vs
        :param policy: name of policy
        :param vs_name: vs name
        :param avi_config: avi config dict
        :param tenant_name: tenant
        :param cloud_name: cloud
        :return: cloned policy object
        """

        policy_name = policy['name']
        clone_policy = copy.deepcopy(policy)
        LOG.debug("cloning policy %s" % clone_policy)
        if 'http_request_policy' in clone_policy:
            for rule in clone_policy['http_request_policy']['rules']:
                if 'switching_action' in rule:
                    if rule['switching_action'].get('pool_group_ref'):
                        pool_group_ref = self.get_name(
                            rule['switching_action']['pool_group_ref'])
                        pool_group_ref = self.clone_pool_group(
                            pool_group_ref, policy_name, avi_config, False,
                            tenant_name, cloud_name)
                        if pool_group_ref:
                            updated_pool_group_ref = self.get_object_ref(
                                pool_group_ref,
                                conv_const.OBJECT_TYPE_POOL_GROUP,
                                tenant_name, cloud_name)
                            rule['switching_action']['pool_group_ref'] = \
                                updated_pool_group_ref
                    elif rule['switching_action'].get('pool_ref'):
                        pool_ref = self.get_name(
                            rule['switching_action']['pool_ref'])
                        if pool_ref:
                            updated_pool_ref = self.get_object_ref(
                                pool_ref, conv_const.OBJECT_TYPE_POOL,
                                tenant_name, cloud_name)
                            rule['switching_action']['pool_ref'] = \
                                updated_pool_ref
        clone_policy['name'] += '-%s-clone' % vs_name
        return clone_policy

    def get_skipped_pool(self, avi_config, pool_name, pool_csv_rows,
                         csv_writer_dict_list, vs_ref, profile_csv_list,
                         skipped_setting):
        """
        This method get the skipped list for pool by going over the
        references attached to it
        :param avi_config: Converted Avi configuration
        :param pool_name: name of the pool
        :param pool_csv_rows:
        :param csv_writer_dict_list: Result report dict
        :param vs_ref: VS reference
        :param profile_csv_list:
        :param skipped_setting: User defined skipped settings
        :return: skipped setting for pool
        """
        pool_skipped_setting = {}
        skipped_list = self.get_pool_skipped(pool_csv_rows, pool_name, vs_ref)
        pool_object = [pool for pool in avi_config["Pool"]
                       if pool['name'] == pool_name]
        if skipped_list:
            pool_skipped_setting['pool_name'] = pool_name
            pool_skipped_setting['pool_skipped_list'] = skipped_list
        if pool_object:
            if 'health_monitor_refs' in pool_object[0]:
                health_monitor_skipped_setting = []
                for health_monitor_ref in pool_object[0]['health_monitor_refs']:
                    health_monitor_ref = self.get_name(health_monitor_ref)
                    monitor_csv_object = self.get_csv_object_list(
                        csv_writer_dict_list, ['monitor'])
                    skipped_list = self.get_csv_skipped_list(
                        monitor_csv_object, health_monitor_ref, vs_ref,
                        field_key='health_monitor')
                    if skipped_list:
                        health_monitor_skipped_setting.append(
                            {'health_monitor_name': health_monitor_ref,
                             'monitor_skipped_list': skipped_list})
                if health_monitor_skipped_setting:
                    pool_skipped_setting['pool_name'] = pool_name
                    pool_skipped_setting['health_monitor'] = \
                        health_monitor_skipped_setting
            if 'ssl_key_and_certificate_ref' in pool_object[0] and \
                    pool_object[0]['ssl_key_and_certificate_ref']:
                ssl_key_cert = self.get_name(
                    pool_object[0]['ssl_key_and_certificate_ref'])
                sslkc_skip = self.get_csv_skipped_list(
                    profile_csv_list, ssl_key_cert, vs_ref,
                    field_key='ssl_cert_key')
                if sslkc_skip:
                    pool_skipped_setting['pool_name'] = pool_name
                    pool_skipped_setting['ssl_key_and_certificate'] = sslkc_skip

            if 'ssl_profile_ref' in pool_object[0] and \
                    pool_object[0]['ssl_profile_ref']:
                name, skipped = self.get_ssl_profile_skipped(
                    profile_csv_list, pool_object[0]['ssl_profile_ref'], vs_ref)
                if skipped:
                    pool_skipped_setting['pool_name'] = pool_name
                    pool_skipped_setting['ssl profile'] = {}
                    pool_skipped_setting['ssl profile']['name'] = name
                    pool_skipped_setting['ssl profile'][
                        'skipped_list'] = skipped

            if 'application_persistence_profile_ref' in pool_object[0] and \
                    pool_object[0]['application_persistence_profile_ref']:
                name, skipped = self.get_app_persistence_profile_skipped(
                    csv_writer_dict_list, pool_object[0], vs_ref)
                if skipped:
                    pool_skipped_setting['pool_name'] = pool_name
                    pool_skipped_setting['Application Persistence profile'] = {}
                    pool_skipped_setting['Application Persistence profile'][
                        'name'] = name
                    pool_skipped_setting['Application Persistence profile'][
                        'skipped_list'] = skipped

            if pool_skipped_setting:
                skipped_setting['pools'].append(pool_skipped_setting)

    def get_skip_pools_policy(self, policy_set_name, http_req, avi_config,
                              pool_csv_rows, vs_ref, profile_csv_list,
                              skipped_setting):
        if http_req['switching_action'].get('pool_group_ref'):
            pool_group_name = self.get_name(http_req['switching_action']
                                            ['pool_group_ref'])
            pool_group_skipped_settings = self.get_pool_skipped_list(
                avi_config, pool_group_name, pool_csv_rows,
                csv_writer_dict_list, vs_ref, profile_csv_list)
            if pool_group_skipped_settings:
                if 'Httppolicy' not in skipped_setting:
                    skipped_setting['Httppolicy'] = {}
                    skipped_setting['Httppolicy']['name'] = policy_set_name
                skipped_setting['Httppolicy']['Pool Group'] =\
                    pool_group_skipped_settings
        elif http_req['switching_action'].get('pool_ref'):
            pool_name = self.get_name(http_req['switching_action']['pool_ref'])
            pool_skipped_settings = {'pools': []}
            self.get_skipped_pool(avi_config, pool_name, pool_csv_rows,
                                  csv_writer_dict_list, vs_ref,
                                  profile_csv_list, pool_skipped_settings)
            if pool_skipped_settings['pools']:
                if 'Httppolicy' not in skipped_setting:
                    skipped_setting['Httppolicy'] = {}
                    skipped_setting['Httppolicy']['name'] = policy_set_name
                skipped_setting['Httppolicy']['Pool'] = pool_skipped_settings

    def remove_pool_group_vrf(self, pool_ref, avi_config):
        """
        This method will remove vrf_ref for all pools in poolgroup
        :param pool_ref: pool group name
        :param avi_config: avi config json
        :return:
        """
        pg_obj = [poolgrp for poolgrp in avi_config['PoolGroup'] if
                  poolgrp['name'] == pool_ref]
        if pg_obj:
            for member in pg_obj[0]['members']:
                poolname = self.get_name(member.get('pool_ref'))
                self.remove_pool_vrf(poolname, avi_config)

    def remove_pool_vrf(self, pool_ref, avi_config):
        """
        This method will remove vrf_ref for pool
        :param pool_ref: pool name
        :param avi_config: avi config json
        :return:
        """
        pool_obj = [pool for pool in avi_config['Pool'] if pool['name'] ==
                    pool_ref]
        if pool_obj and pool_obj[0].get('vrf_ref'):
            pool_obj[0].pop('vrf_ref')
            LOG.debug("Removed vrf ref from the pool %s", pool_ref)

    def update_network_profile(self, aviconfig, sys_dict):
        """
        This method updates the network profile to TCP PROXY when VS has HTTP
        application profile
        :param aviconfig: avi config dict
        :param sys_dict: system config dict
        :return:
        """
        for vs_obj in aviconfig['VirtualService']:
            if vs_obj.get('application_profile_ref'):
                app_profile = self.get_name(vs_obj['application_profile_ref'])
                app_profile_obj = [app for app in sys_dict['ApplicationProfile']
                                   + aviconfig['ApplicationProfile']
                                   if app['name'] == app_profile]
                if app_profile_obj and (
                        app_profile_obj[0]['type'] ==
                        'APPLICATION_PROFILE_TYPE_HTTP' or app_profile_obj[
                        0]['name'] == 'System-HTTP'):
                    if vs_obj.get('network_profile_ref'):
                        nw_profile = self.get_name(vs_obj[
                                                       'network_profile_ref'])
                        nw_profile_obj = [nw for nw in sys_dict[
                                          'NetworkProfile'] + aviconfig[
                                          'NetworkProfile'] if
                                          nw['name'] == nw_profile]
                        if nw_profile_obj and (
                                nw_profile_obj[0]['profile']['type']
                                != 'PROTOCOL_TYPE_TCP_PROXY'):
                            LOG.debug(
                                'Changed the network profile reference from %s '
                                'to TCP-Proxy as VS %s has HTTP profile',
                                nw_profile_obj[0]['profile']['type'],
                                vs_obj['name'])
                            vs_obj['network_profile_ref'] = \
                                self.get_object_ref(
                                    'System-TCP-Proxy',
                                    conv_const.OBJECT_TYPE_NETWORK_PROFILE)

    def correct_vs_ref(self, avi_config):
        """
        This method corrects the reference of VS to different objects
        :param avi_config: avi configuration dict
        :return:
        """
        global csv_writer_dict_list
        avi_graph = self.make_graph(avi_config)
        csv_dict_sub = [row for row in csv_writer_dict_list if row[
                        'F5 type'] != 'virtual' and row['Status'] in
                        (conv_const.STATUS_PARTIAL,
                         conv_const.STATUS_SUCCESSFUL)]
        for dict_row in csv_dict_sub:
            obj = dict_row['Avi Object']
            vs = []
            if obj.startswith('{'):
                obj = eval(obj)
                for key in obj:
                    for objs in obj[key]:
                        self.add_vs_ref(objs, avi_graph, vs)
            elif obj.startswith('['):
                obj = eval(obj)
                for objs in obj:
                    for key in objs:
                        objval = objs[key]
                        self.add_vs_ref(objval, avi_graph, vs)
            if vs:
                dict_row['VS Reference'] = str(list(set(vs)))
            else:
                dict_row['VS Reference'] = conv_const.STATUS_NOT_IN_USE

    def add_vs_ref(self, obj, avi_graph, vs):
        """
        Helper method for adding vs ref
        :param obj: object
        :param avi_graph: avi graph
        :param vs: VS list
        :return:
        """
        tmplist = []

        if isinstance(obj, str) and obj.startswith('Duplicate of'):
            obj_name = None
            LOG.debug("Object has merged: %s" % obj)
        else:
            obj_name = obj.get('name', obj.get('hostname'))
        if obj_name:
            if avi_graph.has_node(obj_name):
                LOG.debug("Checked predecessor for %s", obj_name)
                predecessor = list(avi_graph.predecessors(obj_name))
                if predecessor:
                    self.get_predecessor(predecessor, avi_graph, vs, tmplist)
            else:
                LOG.debug("Object %s may be merged or orphaned", obj_name)

    def get_predecessor(self, predecessor, avi_graph, vs, tmplist):
        """
        This method gets the predecessor of the object
        :param predecessor: predecessor list
        :param avi_graph: avi graph
        :param vs: VS list
        :param tmplist: temporary list of objects for which predecessors
                        are already evaluated
        :return:
        """
        if len(predecessor) > 1:
            for node in predecessor:
                if node in tmplist:
                    continue
                nodelist = [node]
                self.get_predecessor(nodelist, avi_graph, vs, tmplist)
        elif len(predecessor):
            node_obj = [nod for nod in list(avi_graph.nodes().data()) if
                        nod[0] == predecessor[0]]
            if node_obj and (node_obj[0][1]['type'] == 'VS' or 'VS' in node_obj[
              0][1]['type']):
                LOG.debug("Predecessor %s found", predecessor[0])
                vs.extend(predecessor)
            else:
                tmplist.extend(predecessor)
                LOG.debug("Checked predecessor for %s", predecessor[0])
                nodelist = list(avi_graph.predecessors(predecessor[0]))
                self.get_predecessor(nodelist, avi_graph, vs, tmplist)
        else:
            LOG.debug("No more predecessor")

    def convert_irules(self, vs_ds_rules, rule_config, avi_config, prefix,
                       vs_name, tenant):
        vs_ds = list()
        req_policies = list()
        nw_policy = None
        mapped_rules = []
        converted_rules = []

        LOG.debug("Converting for irules %s for vs %s" % (vs_ds_rules, vs_name))

        for rule_mapping in rule_config:
            mapped_rules.append(rule_mapping['rule_name'])

        for index, rule in enumerate(vs_ds_rules):
            rule_mapping = None
            if rule in mapped_rules:
                rule_mapping = [obj for obj in rule_config if
                                obj['rule_name'] == rule][0]
            if rule_mapping and rule_mapping['type'] == 'VSDataScriptSet':
                if 'avi_config' in rule_mapping:
                    ds_config = copy.deepcopy(rule_mapping['avi_config'])
                else:
                    ds_config = copy.deepcopy(conv_const.DUMMY_DS)
                    ds_config['name'] = '%s-dummy' % rule

                ds_config['tenant_ref'] = self.get_object_ref(tenant, 'tenant')
                if prefix:
                    ds_config['name'] = '%s-%s' % (prefix, ds_config['name'])
                existing_ds = [obj for obj in avi_config['VSDataScriptSet']
                               if obj['name'] == ds_config['name']]
                if not existing_ds:
                    avi_config['VSDataScriptSet'].append(ds_config)
                vs_ds.append(ds_config['name'])
                converted_rules.append(rule)
                LOG.debug(
                    "iRule %s successfully mapped to %s VSDataScriptSet" %
                    (rule, ds_config['name']))
            elif rule_mapping and rule_mapping['type'] == 'HTTPPolicySet':
                if 'avi_config' in rule_mapping:
                    policy = copy.deepcopy(rule_mapping['avi_config'])
                    policy['name'] = '%s-%s' % (policy['name'], vs_name)
                else:
                    policy = copy.deepcopy(conv_const.DUMMY_REQ_POLICY)
                    policy['name'] = '%s-%s-dummy' % (rule, vs_name)

                policy['tenant_ref'] = self.get_object_ref(tenant, 'tenant')
                if prefix:
                    policy['name'] = '%s-%s' % (prefix, policy['name'])
                avi_config['HTTPPolicySet'].append(policy)
                req_policies.append(policy['name'])
                converted_rules.append(rule)
                LOG.debug(
                    "iRule %s successfully mapped to %s HTTPPolicySet" %
                    (rule, policy['name']))
            elif rule_mapping and rule_mapping['type'] == \
                    'NetworkSecurityPolicy':
                if 'avi_config' in rule_mapping:
                    policy = copy.deepcopy(rule_mapping['avi_config'])
                    policy['name'] = '%s-%s' % (policy['name'], vs_name)
                else:
                    policy = copy.deepcopy(conv_const.DUMMY_NW_POLICY)
                    policy['name'] = '%s-%s-dummy' % (rule, vs_name)

                policy['tenant_ref'] = self.get_object_ref(tenant, 'tenant')
                if prefix:
                    policy['name'] = '%s-%s' % (prefix, policy['name'])
                avi_config['NetworkSecurityPolicy'].append(policy)
                nw_policy = policy['name']
                converted_rules.append(rule)
                LOG.debug(
                    "iRule %s successfully mapped to %s NetworkSecurityPolicy" %
                    (rule, policy['name']))
            elif (rule_mapping and rule_mapping['type'] ==
                  'HTTPToHTTPSRedirect') or rule == '_sys_https_redirect':
                # Added prefix for objects
                if prefix:
                    policy_name = '%s-%s-%s' % (prefix, rule, vs_name)
                else:
                    policy_name = '%s-%s' % (rule, vs_name)
                policy = copy.deepcopy(conv_const.HTTP_TO_HTTPS_REDIRECT_POL)
                policy["name"] = policy_name
                policy['tenant_ref'] = self.get_object_ref(tenant, 'tenant')
                req_policies.append(policy_name)
                avi_config['HTTPPolicySet'].append(policy)
                converted_rules.append(rule)
                LOG.debug(
                    "iRule %s successfully mapped to %s HTTPPolicySet" %
                    (rule, policy['name']))
        return vs_ds, req_policies, nw_policy, converted_rules

    def update_with_default_profile(self, profile_type, profile,
                                    profile_config, profile_name):
        """
        Profiles can have inheritance used by attribute defaults-from in F5
        configuration this method recursively gets all the attributes from the
        default objects and forms complete object
        :param profile_type: type of profile
        :param profile: currant profile object
        :param profile_config: F5 profile config dict
        :param profile_name: Name of profile
        :return: Complete profile with updated attributes from defaults
        """
        parent_name = profile.get('defaults-from', None)
        if parent_name and profile_name != parent_name:
            parent_profile = profile_config.get(profile_type + " " +
                                                parent_name, None)
            if parent_profile:
                parent_profile = self.update_with_default_profile(
                    profile_type, parent_profile, profile_config, parent_name)
                parent_profile = copy.deepcopy(parent_profile)
                parent_profile.update(profile)
                profile = parent_profile
        return profile