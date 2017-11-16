import csv
import logging
import os
import copy
import re
import random
import urlparse
import ast
import pandas
import pexpect
import avi.migrationtools.netscaler_converter.ns_constants as ns_constants
from pkg_resources import parse_version
from xlsxwriter import Workbook
from openpyxl import load_workbook
from OpenSSL import crypto
from socket import gethostname
from avi.migrationtools.netscaler_converter.ns_constants \
    import (STATUS_SKIPPED, STATUS_SUCCESSFUL, STATUS_INDIRECT,
            STATUS_NOT_APPLICABLE, STATUS_PARTIAL, STATUS_DATASCRIPT,
            STATUS_INCOMPLETE_CONFIGURATION, STATUS_COMMAND_NOT_SUPPORTED,
            OBJECT_TYPE_POOL_GROUP, OBJECT_TYPE_POOL, STATUS_NOT_IN_USE,
            OBJECT_TYPE_HTTP_POLICY_SET, STATUS_LIST, COMPLEXITY_ADVANCED,
            COMPLEXITY_BASIC, OBJECT_TYPE_APPLICATION_PERSISTENCE_PROFILE,
            OBJECT_TYPE_APPLICATION_PROFILE)
from avi.migrationtools.avi_migration_utils import MigrationUtil

LOG = logging.getLogger(__name__)

csv_writer_dict_list = []
skipped_setting = {
    # 'virtual_service': '',
    # 'ssl key and cert': {},
    # 'ssl profile': {},
    # 'pool group': {},
    # 'health monitor': {},
    # 'Httppolicy': {}
}
# Added variable for checking progress and get overall object.
progressbar_count = 0
total_count = 0


class NsUtil(MigrationUtil):

    def add_conv_status(self, line_no, cmd, object_type, full_command, conv_status,
                        avi_object=None):
        """
        Adds as status row in conversion status csv
        :param line_no: line number of command
        :param object_type:
        :param full_command: netscaler command
        :param conv_status: dict of conversion status
        :param avi_object: Converted objectconverted avi object
        """

        row = {
            'Line Number': line_no if line_no else '',
            'Netscaler Command': cmd if cmd else '',
            'Object Name': object_type if object_type else '',
            'Full Command': full_command if full_command else '',
            'Status': conv_status.get('status', ''),
            'Skipped settings': str(conv_status.get('skipped', '')),
            'Indirect mapping': str(conv_status.get('indirect', '')),
            'Not Applicable': str(conv_status.get('na_list', '')),
            'User Ignored': str(conv_status.get('user_ignore', '')),
            'AVI Object': str(avi_object) if avi_object else ''
        }
        csv_writer_dict_list.append(row)

    def add_complete_conv_status(self, ns_config, output_dir, avi_config,
                                 report_name, vs_level_status):
        """
        Adds as status row in conversion status csv
        :param ns_config: NS config dict
        :param output_dir: output directory
        :param avi_config: AVI config dict
        :param report_name: name of report
        :param vs_level_status: add vs level details in XL sheet
        """

        global csv_writer_dict_list
        global progressbar_count
        global total_count
        print "Generating Report For Converted Configuration..."
        ptotal = len(ns_config)
        ppcount = 0
        for config_key in ns_config:
            # increment progressbar count
            ppcount += 1
            config_object = ns_config[config_key]
            msg = "Generating report"
            self.print_progress_bar(ppcount, ptotal, msg, prefix='Progress',
                                    suffix='')
            for element_key in config_object:
                element_object_list = config_object[element_key]
                if isinstance(element_object_list, dict):
                    element_object_list = [element_object_list]
                for element_object in element_object_list:
                    match = [match for match in csv_writer_dict_list
                             if
                             match['Line Number'] == element_object['line_no']]
                    if not match:
                        ns_complete_command = self.get_netscalar_full_command(
                            config_key, element_object)
                        # Add status incomplete configuration
                        self.add_status_row(
                            element_object['line_no'], config_key,
                            element_object['attrs'][0], ns_complete_command,
                            STATUS_INCOMPLETE_CONFIGURATION)
        unique_line_number_list = set()
        row_list = []
        for dict_row in csv_writer_dict_list:
            if dict_row['Line Number'] not in unique_line_number_list:
                unique_line_number_list.add(dict_row['Line Number'])
                row_list.append(dict_row)
            else:
                row = [row for row in row_list
                       if row['Line Number'] == dict_row['Line Number']]
                if str(dict_row['AVI Object']).startswith('Skipped'):
                    continue
                if dict_row.get('AVI Object', None):
                    # Added condition to check unique status.
                    if str(row[0]['AVI Object']) != str(dict_row['AVI Object']):
                        row[0]['AVI Object'] += '__/__%s' % dict_row[
                            'AVI Object']
        for status in STATUS_LIST:
            status_list = [row for row in row_list if
                           row['Status'] == status]
            print '%s: %s' % (status, len(status_list))
        # add skipped list of each object at vs level
        print "Writing Excel Sheet For Converted Configuration..."
        total_count = total_count + len(row_list)
        if vs_level_status:
            self.vs_per_skipped_setting_for_references(avi_config)
            self.correct_vs_ref(avi_config)
        else:
            # Call to calculate vs complexity
            self.vs_complexity_level()
        # Write status report and pivot table in xlsx report
        self.write_status_report_and_pivot_table_in_xlsx(
            row_list, output_dir, report_name, vs_level_status)

    def add_status_row(self, line_no, cmd, object_type, full_command, status,
                       avi_object=None):
        """
        Adds as status row in conversion status csv
        :param line_no:
        :param cmd: netscaler command
        :param object_type:
        :param full_command:
        :param status: conversion status
        :param avi_object:
        """
        global csv_writer_dict_list
        row = {
            'Line Number': line_no if line_no else '',
            'Netscaler Command': cmd,
            'Object Name': object_type,
            'Full Command': full_command,
            'Status': status,
            'AVI Object': str(avi_object) if avi_object else ''
        }
        csv_writer_dict_list.append(row)

    def add_csv_headers(self, csv_file):
        """
        Adds header line in conversion status file
        :param csv_file: File to which header is to be added
        """

        global csv_writer
        fieldnames = ['Line Number', 'Netscaler Command', 'Object Name',
                      'Full Command', 'Status', 'Skipped settings',
                      'Indirect mapping', 'Not Applicable', 'User Ignored',
                      'AVI Object']
        csv_writer = csv.DictWriter(csv_file, fieldnames=fieldnames,
                                    lineterminator='\n', )

        csv_writer.writeheader()

    def get_avi_lb_algorithm(self, ns_algorithm):
        """
        Converts NS LB algorithm to equivalent avi LB algorithm
        :param ns_algorithm: NS algorithm name
        :return: Avi LB algorithm enum value
        """

        avi_algorithm = 'LB_ALGORITHM_LEAST_CONNECTIONS'
        if ns_algorithm == 'LEASTCONNECTIONS':
            avi_algorithm = 'LB_ALGORITHM_LEAST_CONNECTIONS'
        elif ns_algorithm == 'ROUNDROBIN':
            avi_algorithm = 'LB_ALGORITHM_ROUND_ROBIN'
        elif ns_algorithm in ['LEASTRESPONSETIME', 'LRTM']:
            avi_algorithm = 'LB_ALGORITHM_FASTEST_RESPONSE'
        elif ns_algorithm == 'SOURCEIPHASH':
            avi_algorithm = 'LB_ALGORITHM_CONSISTENT_HASH'
        elif ns_algorithm == 'URLHASH':
            avi_algorithm = 'LB_ALGORITHM_CONSISTENT_HASH_URI'
        return avi_algorithm

    def get_avi_resp_code(self, respCode):
        """
        This function used for getting appropriate response code for avi.
        :param respCode: response code
        :return: returns list of unique responses.
        """

        avi_resp_codes = []
        codes = []
        for res_code in respCode.split(' '):
            if '-' in res_code:
                codes.extend(res_code.split('-'))
            else:
                codes.append(res_code)
        for code in codes:
            if code and code.strip().isdigit():
                # Converted to int.
                code = int(code.strip())
                if code < 200:
                    avi_resp_codes.append("HTTP_1XX")
                elif code < 300:
                    avi_resp_codes.append("HTTP_2XX")
                elif code < 400:
                    avi_resp_codes.append("HTTP_3XX")
                elif code < 500:
                    avi_resp_codes.append("HTTP_4XX")
                elif code < 600:
                    avi_resp_codes.append("HTTP_5XX")
        # Get the unique dict from list.
        return list(set(avi_resp_codes))

    def get_conv_status(self, ns_object, skipped_list, na_list, indirect_list,
                        ignore_for_val=None, indirect_commands=None,
                        user_ignore_val=[]):
        """
        This function used for getting status detail for command like
        skipped or indirect.
        :param ns_object: Netscaler parsed config
        :param skipped_list: list of skipped commands list.
        :param na_list: not applicable commands list.
        :param indirect_list: indirect command list
        :param ignore_for_val: optional field
        :param indirect_commands: indirect commands
        :param user_ignore_val: List of user ignore attributes
        :return: returns dict of coversion status.
        """

        skipped = [attr for attr in ns_object.keys() if attr in skipped_list]
        na = [attr for attr in ns_object.keys() if attr in na_list]
        indirect = [attr for attr in ns_object.keys() if attr in indirect_list]
        # List of ignore attributes which are present in skipped
        user_ignore = [val for val in skipped if val in user_ignore_val]
        # Removed the attributes from skipped which are in user ignore list
        skipped = [attr for attr in skipped if attr not in user_ignore_val]
        if ignore_for_val:
            for key in ignore_for_val.keys():
                if key not in ns_object:
                    continue
                ns_val = ns_object.get(key)
                ignore_val = ignore_for_val.get(key)
                if key in skipped and str(ns_val) == str(ignore_val):
                    skipped.remove(key)
        if skipped:
            status = STATUS_PARTIAL
        else:
            status = STATUS_SUCCESSFUL

        conv_status = {
            'skipped': skipped,
            'indirect': indirect,
            'na_list': na,
            'status': status,
            'user_ignore': user_ignore
        }
        return conv_status

    def get_key_cert_obj(self, name, key_file_name, cert_file_name, input_dir):
        """
        :param name:name of ssl cert.
        :param key_file_name:  key file (ie.pem)
        :param cert_file_name: certificate file name
        :param input_dir: input directory for certificate file name
        :return: returns dict of ssl object
        """
        folder_path = input_dir + os.path.sep
        key = self.upload_file(folder_path + key_file_name)
        cert = self.upload_file(folder_path + cert_file_name)
        ssl_kc_obj = None
        if key and cert:
            cert = {"certificate": cert}
            ssl_kc_obj = {
                'name': name,
                'key': key,
                'certificate': cert,
                'key_passphrase': ''
            }
        return ssl_kc_obj

    def get_command_from_line(self, line):
        """
        This function is used for getting command and line number from conf file.
        :param line: line
        :return: returns command name and line
        """

        cmd = ''
        line_no = 0
        for member in line:
            if 'line_no' in member:
                line_no = member[1]
                continue
            if isinstance(member, str):
                cmd += ' %s' % member
            else:
                cmd += ' -%s' % ' '.join(member)
        return cmd, line_no

    def update_status_for_skipped(self, skipped_cmds):
        """
        :param skipped_cmds: separation of non converted commands
         to NA, Indirect,DataScript,NotSupported
        :return: None
        """

        na_cmds = ns_constants.netscalar_command_status['NotApplicableCommands']
        indirect_cmds = ns_constants.netscalar_command_status[
            'IndirectCommands']
        datascript_cmds = \
            ns_constants.netscalar_command_status['DatascriptCommands']
        not_supported = ns_constants.netscalar_command_status['NotSupported']
        if not skipped_cmds:
            return
        for cmd in skipped_cmds:
            line_no = cmd['line_no']
            cmd = cmd['cmd']
            cmd = cmd.strip()
            for na_cmd in na_cmds:
                if cmd.startswith(na_cmd):
                    # Add status not applicable in csv/report
                    self.add_status_row(line_no, na_cmd, None, cmd,
                                   STATUS_NOT_APPLICABLE)
                    break
            for id_cmd in indirect_cmds:
                if cmd.startswith(id_cmd):
                    # Add status indirect in csv/report
                    self.add_status_row(line_no, id_cmd, None, cmd, STATUS_INDIRECT)
                    break
            for datascript_cmd in datascript_cmds:
                if cmd.startswith(datascript_cmd):
                    # Add status datascript in csv/report
                    self.add_status_row(line_no, datascript_cmd, None, cmd,
                                   STATUS_DATASCRIPT)
                    break
            for not_commands in not_supported:
                if cmd.startswith(not_commands):
                    # Add status not not supported in csv/report
                    self.add_status_row(line_no, not_commands, None, cmd,
                                   STATUS_COMMAND_NOT_SUPPORTED)
                    break

    def remove_duplicate_objects(self, obj_type, obj_list):
        """
        Remove duplicate objects from list
        :param obj_type: Object type
        :param obj_list: list of all objects
        :return: return list which has no duplicates objects
        """

        if len(obj_list) == 1:
            return obj_list
        for source_obj in obj_list:
            for index, tmp_obj in enumerate(obj_list):
                if tmp_obj["name"] == source_obj["name"]:
                    continue
                src_cp = copy.deepcopy(source_obj)
                tmp_cp = copy.deepcopy(tmp_obj)
                del src_cp["name"]
                if "description" in src_cp:
                    del src_cp["description"]

                del tmp_cp["name"]
                if "description" in tmp_cp:
                    del tmp_cp["description"]
                if cmp(src_cp, tmp_cp) == 0:
                    LOG.warn('Remove duplicate %s object : %s' % (obj_type,
                                                                  tmp_obj[
                                                                      "name"]))
                    del obj_list[index]
                    self.remove_duplicate_objects(obj_type, obj_list)
        return obj_list

    def cleanup_config(self, config):
        """
        This function is used for deleting temp variables created for conversion
        :param config: dict type
        :return: None
        """

        del config

    def clone_pool(self, pool_name, cloned_for, avi_config, userprefix=None):
        """
        This function used for cloning shared pools in netscaler.
        :param pool_name: name of pool
        :param cloned_for: cloned for
        :param avi_config: avi config dict
        :param userprefix: prefix for objects
        :return: None
        """
        pools = [pool for pool in avi_config['Pool'] if
                 pool['name'] == pool_name]
        if pools:
            pool_obj = copy.deepcopy(pools[0])
            pname = pool_obj['name']
            pool_name = re.sub('[:]', '-', '%s-%s' % (pname, cloned_for))
            pool_obj['name'] = pool_name
            avi_config['Pool'].append(pool_obj)
            LOG.info(
                "Same pool reference to other object. Clone Pool %s for %s" %
                (pool_name, cloned_for))
            return pool_obj['name']
        return None

    def get_vs_if_shared_vip(self, avi_config, controller_version):
        """
        This function checks if same vip is used for other vs
        :param avi_config: avi config dict
        :param controller_version:
        :return: None
        """

        vs_list = [v for v in avi_config['VirtualService'] if
                   'port_range_end' in
                   v['services'][0]]
        for vs in vs_list:
            # Get the list of vs which shared the same vip
            if parse_version(controller_version) >= parse_version('17.1'):
                vs_port_list = [int(v['services'][0]['port']) for v in
                                avi_config['VirtualService']
                                if v['vip'][0]['ip_address']['addr'] ==
                                vs['vip'][0]['ip_address']['addr']
                                and 'port_range_end' not in v['services'][0]]
            else:
                vs_port_list = [int(v['services'][0]['port']) for v in
                                avi_config['VirtualService'] if v['ip_address'][
                                    'addr'] == vs['ip_address']['addr'] and
                                'port_range_end' not in v['services'][0]]

            if vs_port_list:
                min_port = min(vs_port_list)
                max_port = max(vs_port_list)
                vs['services'][0]['port_range_end'] = str(min_port - 1)
                service = {
                    'enable_ssl': False,
                    'port': str(max_port + 1),
                    'port_range_end': '65535'
                }
                vs['services'].append(service)

    def add_prop_for_http_profile(self, profile_name, avi_config, sysdict,
                                  prop_dict):
        """
        This method adds the additional attribute to application profile
        :param profile_name: name of application profile
        :param avi_config: avi config dict
        :param sysdict: system/baseline config dict
        :param prop_dict: property dict
        :return:
        """

        profile = [p for p in (avi_config['ApplicationProfile'] + sysdict[
            'ApplicationProfile']) if p['name'] == profile_name]
        if profile:
            if prop_dict.get('clttimeout'):
                profile[0]['client_header_timeout'] = int(prop_dict[
                                                              'clttimeout'])
                profile[0]['client_body_timeout'] = int(prop_dict['clttimeout'])
            if prop_dict.get('xff_enabled'):
                if profile[0].get('http_profile'):
                    profile[0]['http_profile'].update(
                        {
                            'xff_enabled': True,
                            'xff_alternate_name': 'X-Forwarded-For'
                        }
                    )
                else:
                    profile[0].update({'http_profile':
                        {
                            'xff_enabled': True,
                            'xff_alternate_name': 'X-Forwarded-For'
                        }
                    })
            if prop_dict.get('ssl_everywhere_enabled'):
                if profile[0].get('http_profile'):
                    profile[0]['http_profile'].update(
                        {
                            'ssl_everywhere_enabled': True,
                            'x_forwarded_proto_enabled': True,
                            'hsts_enabled': True,
                            'http_to_https': True,
                            'httponly_enabled': True,
                            'hsts_max_age': 365,
                            'server_side_redirect_to_https': True,
                            'secure_cookie_enabled': True
                        }
                    )
                else:
                    profile[0].update({'http_profile':
                        {
                            'ssl_everywhere_enabled': True,
                            'x_forwarded_proto_enabled': True,
                            'hsts_enabled': True,
                            'http_to_https': True,
                            'httponly_enabled': True,
                            'hsts_max_age': 365,
                            'server_side_redirect_to_https': True,
                            'secure_cookie_enabled': True
                        }
                    })

    def object_exist(self, object_type, name, avi_config):
        '''
        This method returns true if object exists in avi config dict else false
        :param object_type:
        :param name:
        :param avi_config:
        :return:
        '''
        data = avi_config[object_type]
        obj_list = [obj for obj in data if obj['name'] == name]
        if obj_list:
            return True
        return False

    def is_shared_same_vip(self, vs, cs_vs_list, avi_config, tenant_name,
                           cloud_name, tenant_ref, cloud_ref,
                           controller_version, prefix):
        """
        This function check for vs sharing same vip
        :param vs: Name of vs
        :param cs_vs_list: List of vs
        :param avi_config: avi config dict
        :param tenant_name: Name of tenant
        :param cloud_name: Name of cloud
        :param tenant_ref: Reference of tenant
        :param cloud_ref: Reference of cloud
        :param controller_version: controller version
        :param prefix: prefix for objects
        :return: None
        """

        if parse_version(controller_version) >= parse_version('17.1'):
            # Get the list of vs which shared the same vip
            shared_vip = [v for v in cs_vs_list if v['vip'][0]['ip_address'][
                'addr'] == vs['vip'][0]['ip_address']['addr'] and
                          v['services'][0][
                              'port'] == vs['services'][0]['port']]
        else:
            shared_vip = [v for v in cs_vs_list if v['ip_address']['addr'] ==
                          vs['ip_address']['addr'] and v['services'][0][
                              'port'] ==
                          vs['services'][0]['port']]

        if shared_vip:
            return True
        elif parse_version(controller_version) >= parse_version('17.1'):
            vsvip = vs['vip'][0]['ip_address']['addr']
            self.create_update_vsvip(vsvip, avi_config['VsVip'], tenant_ref,
                                cloud_ref,
                                prefix=prefix)
            name = vsvip + '-vsvip'
            # Added prefix for objects
            if prefix:
                name = prefix + '-' + vsvip + '-vsvip'
            updated_vsvip_ref = self.get_object_ref(
                name, 'vsvip', tenant_name, cloud_name)
            vs['vsvip_ref'] = updated_vsvip_ref

    def clone_http_policy_set(self, policy, prefix, avi_config, tenant_name,
                              cloud_name, used_poolgrp_ref, userprefix=None):
        """
        This function clone pool reused in context switching rule
        :param policy: name of policy
        :param prefix: clone for
        :param avi_config: avi config dict
        :param tenant_name:
        :param cloud_name:
        :param used_poolgrp_ref:
        :param userprefix: prefix for objects
        :return:None
        """

        policy_name = policy['name']
        clone_policy = copy.deepcopy(policy)
        for rule in clone_policy['http_request_policy']['rules']:
            if rule.get('switching_action', None) and \
                    rule['switching_action'].get('pool_group_ref'):
                pool_group_ref = \
                    rule['switching_action']['pool_group_ref'].split('&')[
                        1].split(
                        '=')[1]
                if pool_group_ref in used_poolgrp_ref:
                    LOG.debug('Cloned the pool group for policy %s',
                              policy_name)
                    pool_group_ref = self.clone_pool_group(
                        pool_group_ref, policy_name, avi_config, tenant_name,
                        cloud_name, userprefix=userprefix)
                if pool_group_ref:
                    updated_pool_group_ref = self.get_object_ref(
                        pool_group_ref, OBJECT_TYPE_POOL_GROUP, tenant_name,
                        cloud_name)
                    rule['switching_action']['pool_group_ref'] = \
                        updated_pool_group_ref
        clone_policy['name'] += '-%s-clone' % prefix
        return clone_policy

    def set_rules_index_for_http_policy_set(self, avi_config):
        """
        Update index as per avi protobuf requirements
        :param avi_config: avi config dict
        :return: None
        """

        http_policy_sets = avi_config['HTTPPolicySet']
        for http_policy_set in http_policy_sets:
            rules = http_policy_set['http_request_policy']['rules']
            rules = sorted(rules, key=lambda d: int(d['index']))
            for index, rule in enumerate(rules):
                rule['index'] = index

    def get_netscalar_full_command(self, netscalar_command, obj):
        """
        Generate netscaler command from the parse dict
        :param netscalar_command: name of command
        :param obj: object with attributes
        :return: Full command
        """

        for attr in obj['attrs']:
            netscalar_command += ' %s' % attr
        for key in obj:
            if isinstance(obj[key], list):
                continue
            if key == 'line_no':
                continue
            netscalar_command += ' -%s %s' % (key, obj[key])
        return netscalar_command

    def clone_pool_group(self, pg_name, cloned_for, avi_config, tenant_name,
                         cloud_name, userprefix=None):
        """
        Used for cloning shared pool group.
        :param pg_name: pool group name
        :param cloned_for: clone for
        :param avi_config: avi config dict
        :param tenant_name:
        :param cloud_name:
        :param userprefix: prefix for objects
        :return: None
        """
        pool_groups = [pg for pg in avi_config['PoolGroup']
                       if pg['name'] == pg_name]
        if pool_groups:
            pool_group = copy.deepcopy(pool_groups[0])
            pool_group_name = re.sub('[:]', '-',
                                     '%s-%s' % (pg_name, cloned_for))
            pool_group['name'] = pool_group_name
            for member in pool_group.get('members', []):
                pool_ref = self.get_name(member['pool_ref'])
                pool_ref = self.clone_pool(pool_ref, cloned_for, avi_config,
                                      userprefix=userprefix)
                if pool_ref:
                    updated_pool_ref = self.get_object_ref(
                        pool_ref, OBJECT_TYPE_POOL, tenant_name, cloud_name)
                    member['pool_ref'] = updated_pool_ref
            avi_config['PoolGroup'].append(pool_group)
            LOG.info(
                "Same pool group reference to other object. Clone Pool group "
                "%s for %s" % (pg_name, cloned_for))
            return pool_group['name']
        return None

    def remove_http_mon_from_pool(self, avi_config, pool, sysdict):
        """
        This function is used for removing http type health monitor from https
        vs.
        :param avi_config: avi config dict
        :param pool: name of pool
        :param sysdict: baseline/system config dict
        :return: None
        """
        if pool:
            hm_refs = copy.deepcopy(pool['health_monitor_refs'])
            for hm_ref in hm_refs:
                hm = [h for h in (sysdict['HealthMonitor'] + avi_config[
                    'HealthMonitor']) if h['name'] == hm_ref]
                if hm and hm[0]['type'] == 'HEALTH_MONITOR_HTTP':
                    pool['health_monitor_refs'].remove(hm_ref)
                    LOG.warning(
                        'Skipping %s this reference from %s pool because '
                        'of health monitor type is HTTP and VS has ssl '
                        'profile.' % (hm_ref, pool['name']))

    def remove_https_mon_from_pool(self, avi_config, pool, sysdict):
        """
        This function is used for removing https type health monitor from http
        vs.
        :param avi_config: avi config dict
        :param pool: name of pool
        :param sysdict: baseline/system config dict
        :return: None
        """
        if pool:
            hm_refs = copy.deepcopy(pool['health_monitor_refs'])
            for hm_ref in hm_refs:
                hm = [h for h in (sysdict['HealthMonitor'] + avi_config[
                    'HealthMonitor']) if h['name'] == hm_ref]
                if hm and hm[0]['type'] == 'HEALTH_MONITOR_HTTPS':
                    pool['health_monitor_refs'].remove(hm_ref)
                    LOG.warning(
                        'Skipping %s this reference from %s pool because '
                        'of health monitor type is HTTPS and VS has no ssl '
                        'profile.' % (hm_ref, pool['name']))

    def update_application_profile(self, profile_name, pki_profile_ref,
                                   tenant_ref, name, avi_config, sysdict):
        """
        This functions defines to update application profile with pki profile if
        application profile exist if not create new http profile with pki profile
        :param profile_name: name of Http profile
        :param pki_profile_ref:  ref of PKI profile
        :param tenant_ref: tenant ref
        :param name: name of virtual service
        :param avi_config: Dict of AVi config
        :param sysdict: baseline/system config
        :return: Http profile
        """

        try:
            if profile_name:
                app_profile = [p for p in (sysdict['ApplicationProfile'] +
                                           avi_config['ApplicationProfile']) if
                               p['name'] ==
                               profile_name]
                if app_profile:
                    app_profile[0]["http_profile"]['pki_profile_ref'] = \
                        pki_profile_ref
                    LOG.debug('Added PKI profile to application profile '
                              'successfully : %s' % (
                              profile_name, pki_profile_ref))
            else:
                app_profile = dict()
                app_profile['name'] = name + '-%s-dummy' % random.randrange(0,
                                                                            1000)
                app_profile['tenant_ref'] = tenant_ref
                app_profile['type'] = 'APPLICATION_PROFILE_TYPE_HTTP'
                http_profile = dict()
                http_profile['connection_multiplexing_enabled'] = False
                http_profile['xff_enabled'] = False
                # TODO: clientIpHdrExpr conversion to xff_alternate_name
                http_profile['websockets_enabled'] = False
                http_profile['pki_profile_ref'] = pki_profile_ref
                app_profile["http_profile"] = http_profile
                avi_config['ApplicationProfile'].append(app_profile)
                LOG.debug(
                    "Conversion completed successfully for httpProfile: %s" %
                    app_profile['name'])
                return app_profile['name']
        except:
            LOG.error("Error in convertion of httpProfile", exc_info=True)

    def convert_persistance_prof(self, vs, name, tenant_ref):
        """
        This function defines that it convert the persistent profile and
        return that profile
        :param vs: object of lb vs or pool
        :param name: name of application persteance profile
        :param tenant_ref: reference of tenant
        :return: application persistent profile
        """

        profile = None
        persistenceType = vs.get('persistenceType', '')
        if persistenceType == 'COOKIEINSERT':
            timeout = vs.get('timeout', 2)
            profile = {
                "http_cookie_persistence_profile": {
                    "always_send_cookie": False
                },
                "persistence_type": "PERSISTENCE_TYPE_HTTP_COOKIE",
                "server_hm_down_recovery": "HM_DOWN_PICK_NEW_SERVER",
                "name": name,
            }
            #  Added time if greater than zero
            if int(timeout) > 0:
                profile['http_cookie_persistence_profile']["timeout"] = timeout
        elif persistenceType == 'SOURCEIP':
            # Set timeout equal to 2 if not provided.
            timeout = vs.get('timeout', 120)
            timeout = int(timeout) / 60
            if timeout < 1:
                timeout = 1
            profile = {
                "server_hm_down_recovery": "HM_DOWN_PICK_NEW_SERVER",
                "persistence_type": "PERSISTENCE_TYPE_CLIENT_IP_ADDRESS",
                "ip_persistence_profile": {
                    "ip_persistent_timeout": timeout
                },
                "name": name
            }
        elif persistenceType == 'SSLSESSION':
            profile = {
                "server_hm_down_recovery": "HM_DOWN_PICK_NEW_SERVER",
                "persistence_type": "PERSISTENCE_TYPE_TLS",
                "name": name
            }
        profile['tenant_ref'] = tenant_ref
        return profile

    def update_status_target_lb_vs_to_indirect(self, larget_lb_vs):
        """
        This function defines that update status for the target lb vserver as
        Indirect
        :param larget_lb_vs: name of target lb vserver
        :return: None
        """
        global csv_writer_dict_list
        row = [row for row in csv_writer_dict_list
               if row['Object Name'] == larget_lb_vs
               and row['Netscaler Command'] == 'add lb vserver']
        if row:
            row[0]['Status'] = STATUS_INDIRECT

    def create_http_policy_set_for_redirect_url(self, vs_obj, redirect_uri,
                            avi_config, tenant_name, tenant_ref, enable_ssl):
        """
        This function defines that create http policy for redirect url
        :param vs_obj: object of VS
        :param redirect_uri: redirect uri
        :param avi_config: dict of AVi
        :param tenant_name: name of tenant
        :param tenant_ref: tenant ref
        :param enable_ssl: flag for enabling ssl
        :return: None
        """
        redirect_uri = str(redirect_uri).replace('"', '')
        action = self.build_redirect_action_dict(redirect_uri, enable_ssl)
        policy_obj = {
            'name': vs_obj['name'] + '-redirect-policy',
            'tenant_ref': tenant_ref,
            'enable': 'false',
            'http_request_policy': {
                'rules': [
                    {
                        'index': 0,
                        'name': vs_obj['name'] + '-redirect-policy-rule-0',
                        'match': {
                            'path': {
                                'match_case': 'INSENSITIVE',
                                'match_str': [
                                    '/'
                                ],
                                'match_criteria': 'EQUALS'
                            }
                        },
                        'redirect_action': action
                    }
                ]
            }
        }
        updated_http_policy_ref = self.get_object_ref(policy_obj['name'],
                                                 OBJECT_TYPE_HTTP_POLICY_SET,
                                                 tenant_name)
        http_policies = {
            'index': 11,
            'http_policy_set_ref': updated_http_policy_ref
        }
        if not vs_obj.get('http_policies'):
            vs_obj['http_policies'] = []
        else:
            ind = max([policies['index'] for policies in vs_obj[
                'http_policies']])
            http_policies['index'] = ind + 1
        vs_obj['http_policies'].append(http_policies)
        avi_config['HTTPPolicySet'].append(policy_obj)

    def clean_virtual_service_from_avi_config(self, avi_config,
                                              controller_version):
        """
        This function defines that clean up vs which has vip 0.0.0.0
        :param avi_config: dict of AVI
        :param controller_version:
        :return: None
        """
        vs_list = copy.deepcopy(avi_config['VirtualService'])
        avi_config['VirtualService'] = []
        if parse_version(controller_version) >= parse_version('17.1'):
            avi_config['VirtualService'] = \
                [vs for vs in vs_list
                 if vs['vip'][0]['ip_address']['addr'] != '0.0.0.0']
        else:
            avi_config['VirtualService'] = \
                [vs for vs in vs_list
                 if vs['ip_address']['addr'] != '0.0.0.0']

    def parse_url(self, url):
        """
        This method returns the parsed url
        :param url: url that need to be parsed
        :return:
        """
        parsed = urlparse.urlparse(url)
        return parsed

    def format_string_to_json(self, avi_string):
        """
        This function defines that it convert string into json format to
        convert into dict
        :param avi_string: string to be converted
        :return: Return converted string
        """
        avi_string = avi_string.split('__/__')[0]
        return ast.literal_eval(avi_string)

    def get_csv_object_list(self, csv_writer_dict_list, command_list):
        """
        This method is used for getting csv object
        :param csv_writer_dict_list: CSV row of object from xlsx report
        :param command_list: List of netscaler commands
        :return: List of CSV rows
        """
        csv_object = [row for row in
                      csv_writer_dict_list
                      if row['Status'] in [STATUS_PARTIAL, STATUS_SUCCESSFUL]
                      and row['Netscaler Command'] in
                      command_list]
        return csv_object

    def get_csv_skipped_list(self, csv_object, name_of_object, vs_ref):
        """
        This method is used for getting skipped list from vs.
        :param csv_object: CSV row of object from xlsx report
        :param name_of_object: Name of Object
        :param vs_ref: Reference of VS
        :return: List of skipped settings
        """

        skipped_list = []
        for each_partial in csv_object:
            avi_object_json = \
                self.format_string_to_json(each_partial['AVI Object'])
            if avi_object_json.get('name') and \
                            avi_object_json['name'] == name_of_object:
                # Set the VS reference for Netscaler status row
                each_partial['VS Reference'] = vs_ref
                repls = ('[', ''), (']', '')
                skipped_setting_csv = reduce(lambda a, kv: a.replace(*kv),
                                             repls,
                                             each_partial['Skipped settings'])
                if skipped_setting_csv:
                    skipped_list.append(skipped_setting_csv)
        return skipped_list

    def get_ssl_key_and_cert_refs_skipped(self, csv_writer_dict_list,
                                          object_name, vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param csv_writer_dict_list: CSV row of object from xlsx report
        :param object_name: like virtual service or pool name
        :param vs_ref: Reference of VS
        :return: List of skipped settings
        """

        ssl_key_cert = \
            self.get_name(object_name['ssl_key_and_certificate_refs'][0])
        csv_object = self.get_csv_object_list(
            csv_writer_dict_list, ['bind ssl vserver', 'bind ssl service',
                                   'bind ssl serviceGroup'])
        skipped_list = self.get_csv_skipped_list(csv_object, ssl_key_cert,
                                                 vs_ref)
        return ssl_key_cert, skipped_list

    def get_ssl_profile_skipped(self, csv_writer_dict_list, ssl_profile_ref,
                                vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param csv_writer_dict_list: CSV row of object from xlsx report
        :param ssl_profile_ref: reference of ssl profile object
        :param vs_ref: virtual service obj reference.
        :return: List of skipped settings
        """

        ssl_profile_name = self.get_name(ssl_profile_ref)
        csv_object = \
            self.get_csv_object_list(csv_writer_dict_list,
                                ['set ssl vserver', 'set ssl service',
                                 'set ssl serviceGroup'])
        skipped_list = self.get_csv_skipped_list(csv_object, ssl_profile_name,
                                            vs_ref)
        return ssl_profile_name, skipped_list

    def get_application_profile_skipped(self, csv_writer_dict_list,
                                        name_of_object, vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param csv_writer_dict_list: CSV row of object from xlsx report
        :param name_of_object: object name like pool name, etc
        :param vs_ref: virtual service obj reference.
        :return: List of skipped settings
        """

        ssl_profile_name = self.get_name(
            name_of_object['application_profile_ref'])
        csv_object = self.get_csv_object_list(
            csv_writer_dict_list, ['add ns httpProfile'])
        skipped_list = self.get_csv_skipped_list(csv_object, ssl_profile_name,
                                                 vs_ref)
        return ssl_profile_name, skipped_list

    def get_network_profile_skipped(self, csv_writer_dict_list, name_of_object,
                                    vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param csv_writer_dict_list:List of add ns tcpProfile netscaler command rows
        :param name_of_object: object name like pool name, etc
        :param vs_ref: virtual service obj reference.
        :return: List of skipped settings
        """

        ssl_profile_name = self.get_name(name_of_object['network_profile_ref'])
        csv_object = self.get_csv_object_list(
            csv_writer_dict_list, ['add ns tcpProfile'])
        skipped_list = self.get_csv_skipped_list(csv_object, ssl_profile_name,
                                            vs_ref)
        return ssl_profile_name, skipped_list

    def get_app_persistence_profile_skipped(self, csv_writer_dict_list,
                                            name_of_object, vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param csv_writer_dict_list: List of set lb group netscaler command rows
        :param name_of_object: object name like pool name, etc
        :param vs_ref: virtual service obj reference.
        :return: List of skipped settings
        """
        # Changed ssl profile name to ssl profile ref.
        app_persistence_profile_name = self.get_name(
            name_of_object['ssl_profile_ref'])
        csv_object = self.get_csv_object_list(csv_writer_dict_list, ['set lb group'])
        skipped_list = self.get_csv_skipped_list(
            csv_object, app_persistence_profile_name, vs_ref)
        return app_persistence_profile_name, skipped_list

    def get_pool_skipped_list(self, avi_config, pool_group_name,
                              skipped_setting, csv_object, obj_name,
                              csv_writer_dict_list, vs_ref):
        """
        This method is used for getting pool skipped list.
        :param avi_config: AVI dict
        :param pool_group_name: Name of Pool group
        :param skipped_setting: List of skipped settings
        :param csv_object: CSV row
        :param obj_name: Name of Object
        :param csv_writer_dict_list: List of bind lb vserver netscaler command
                                     rows
        :param vs_ref: vs object reference
        :return: List of skipped settings
        """

        pool_group_object_ref = [pool_group_object_ref for pool_group_object_ref
                                 in avi_config['PoolGroup'] if
                                 pool_group_object_ref[
                                     'name'] == pool_group_name]
        for pool_group in pool_group_object_ref:
            if 'members' in pool_group:
                for each_pool_ref in pool_group['members']:
                    pool_name = self.get_name(each_pool_ref['pool_ref'])
                    skipped_list = self.get_csv_skipped_list(csv_object, pool_name,
                                                        vs_ref)
                    if len(skipped_list) > 0:
                        skipped_setting[obj_name] = {}
                        skipped_setting[obj_name]['pool'] = {}
                        skipped_setting[obj_name]['pool'][
                            'pool_name'] = pool_name
                        skipped_setting[obj_name]['pool']['pool_skipped_list'] \
                            = skipped_list
                    for pool_partial in csv_object:
                        avi_object_json = self.format_string_to_json(
                            pool_partial['AVI Object'])
                        if avi_object_json['name'] == pool_name:
                            if 'health_monitor_refs' in avi_object_json and \
                                    avi_object_json['health_monitor_refs']:
                                monitor_refs = \
                                    avi_object_json['health_monitor_refs']
                                for monitor_ref in monitor_refs:
                                    monitor_ref = self.get_name(monitor_ref)
                                    csv_object = self.get_csv_object_list(
                                        csv_writer_dict_list,
                                        ['add lb monitor'])
                                    skipped_list = self.get_csv_skipped_list(
                                        csv_object, monitor_ref, vs_ref)
                                    if skipped_list:
                                        skipped_setting[obj_name] = {}
                                        skipped_setting[obj_name]['pool'] = {}
                                        skipped_setting[obj_name]['pool'][
                                            'pool_name'] = pool_name
                                        skipped_setting[obj_name]['pool'][
                                            'health monitor'] = {}
                                        skipped_setting[obj_name]['pool'][
                                            'health monitor'][
                                            'name'] = monitor_ref
                                        skipped_setting[obj_name]['pool'][
                                            'health monitor']['skipped_list'] =\
                                            skipped_list
                            if 'ssl_key_and_certificate_refs' in avi_object_json:
                                name, skipped = \
                                    self.get_ssl_key_and_cert_refs_skipped(
                                        csv_writer_dict_list, avi_object_json,
                                        vs_ref)
                                if skipped:
                                    skipped_setting[obj_name] = {}
                                    skipped_setting[obj_name]['pool'] = {}
                                    skipped_setting[obj_name]['pool'][
                                        'pool_name'] = pool_name
                                    skipped_setting[
                                        obj_name]['pool'][
                                        'ssl key and cert'] = {}
                                    skipped_setting[
                                        obj_name]['pool']['ssl key and cert'][
                                        'name'] = name
                                    skipped_setting[
                                        obj_name]['pool']['ssl key and cert'][
                                        'skipped_list'] = skipped
                            if 'ssl_profile_ref' in avi_object_json:
                                name, skipped = \
                                    self.get_ssl_profile_skipped(
                                        csv_writer_dict_list, avi_object_json[
                                            'ssl_profile_ref'], vs_ref)
                                if skipped:
                                    skipped_setting[obj_name] = {}
                                    skipped_setting[obj_name]['pool'] = {}
                                    skipped_setting[obj_name]['pool'][
                                        'pool_name'] = pool_name
                                    skipped_setting[obj_name]['pool'][
                                        'ssl profile'] = {}
                                    skipped_setting[obj_name]['pool'][
                                        'ssl profile']['name'] = name
                                    skipped_setting[obj_name]['pool'][
                                        'ssl profile']['skipped_list'] = skipped
                            # Get the skipped settings of application
                            # persistence profile ref.
                            if 'application_persistence_profile_ref' in \
                                    avi_object_json:
                                name, skipped = \
                                    self.get_app_persistence_profile_skipped(
                                        csv_writer_dict_list, avi_object_json,
                                        vs_ref)
                                if skipped:
                                    skipped_setting[obj_name] = {}
                                    skipped_setting[obj_name]['pool'] = {}
                                    skipped_setting[obj_name]['pool'][
                                        'pool_name'] = pool_name
                                    skipped_setting[obj_name]['pool'][
                                        'Application Persistence profile'] = {}
                                    skipped_setting[obj_name]['pool'][
                                        'Application Persistence profile'][
                                        'name'] = name
                                    skipped_setting[obj_name]['pool'][
                                        'Application Persistence profile'][
                                        'skipped_list'] = skipped
                            # Get the skipped settings of application
                            # persistence profile ref.
                            if 'application_persistence_profile_ref' \
                                    in avi_object_json:
                                name, skipped = \
                                    self.get_app_persistence_profile_skipped(
                                    csv_writer_dict_list, avi_object_json,
                                    vs_ref)
                                if skipped:
                                    skipped_setting[obj_name] = {}
                                    skipped_setting[obj_name]['pool'] = {}
                                    skipped_setting[obj_name]['pool'][
                                        'pool_name'] = pool_name
                                    skipped_setting[obj_name]['pool'][
                                        'Application Persistence profile'] = {}
                                    skipped_setting[obj_name]['pool'][
                                        'Application Persistence profile'][
                                        'name'] = name
                                    skipped_setting[obj_name]['pool'][
                                        'Application Persistence profile'][
                                        'skipped_list'] = skipped

    def vs_complexity_level(self):
        """
        This method calculate complexity of vs.
        :return:
        """
        vs_csv_objects = [row for row in csv_writer_dict_list
                          if
                          row['Status'] in [STATUS_PARTIAL, STATUS_SUCCESSFUL]
                          and row['Netscaler Command'] in [
                              'add cs vserver', 'add lb vserver']]
        for vs_csv_object in vs_csv_objects:
            virtual_service = self.format_string_to_json(
                vs_csv_object['AVI Object'])
            # Update the complexity level of VS as Basic or Advanced
            self.update_vs_complexity_level(vs_csv_object, virtual_service)

    def vs_per_skipped_setting_for_references(self, avi_config):
        """
        This functions defines that Add the skipped setting per VS CSV row
        :param avi_config: this methode use avi_config for checking vs skipped
        :return: None
        """

        # Get the count of vs sucessfully migrated
        global fully_migrated
        global total_count
        global progressbar_count
        fully_migrated = 0
        # Get the VS object list which is having status successful and partial.
        vs_csv_objects = [row for row in csv_writer_dict_list
                          if
                          row['Status'] in [STATUS_PARTIAL, STATUS_SUCCESSFUL]
                          and row['Netscaler Command'] in [
                              'add cs vserver', 'add lb vserver']]
        # calculate total count
        total_count = total_count + len(vs_csv_objects)
        for vs_csv_object in vs_csv_objects:
            progressbar_count += 1
            skipped_setting = {}
            virtual_service = self.format_string_to_json(
                vs_csv_object['AVI Object'])
            # Update the complexity level of VS as Basic or Advanced
            self.update_vs_complexity_level(vs_csv_object, virtual_service)
            vs_ref = virtual_service['name']
            repls = ('[', ''), (']', '')
            # Get list of skipped setting attributes
            skipped_setting_csv = reduce(lambda a, kv: a.replace(*kv), repls,
                                         vs_csv_object['Skipped settings'])
            if skipped_setting_csv:
                skipped_setting['virtual_service'] = [skipped_setting_csv]
            # Get the skipped list for ssl key and cert
            if 'ssl_key_and_certificate_refs' in virtual_service:
                name, skipped = self.get_ssl_key_and_cert_refs_skipped(
                    csv_writer_dict_list, virtual_service, vs_ref)
                if skipped:
                    skipped_setting['ssl key and cert'] = {}
                    skipped_setting['ssl key and cert']['name'] = name
                    skipped_setting['ssl key and cert'][
                        'skipped_list'] = skipped
            # Get the skipped list for ssl profile name.
            # Changed ssl profile name to ssl profile ref.
            if 'ssl_profile_ref' in virtual_service:
                name, skipped = self.get_ssl_profile_skipped(
                    csv_writer_dict_list, virtual_service['ssl_profile_ref'],
                    vs_ref)
                if skipped:
                    skipped_setting['ssl profile'] = {}
                    skipped_setting['ssl profile']['name'] = name
                    skipped_setting['ssl profile']['skipped_list'] = skipped
            # Get the skipped list for pool group.
            if 'pool_group_ref' in virtual_service:
                pool_group_name = self.get_name(
                    virtual_service['pool_group_ref'])
                csv_object = self.get_csv_object_list(
                    csv_writer_dict_list, ['bind lb vserver'])
                self.get_pool_skipped_list(
                    avi_config, pool_group_name, skipped_setting, csv_object,
                    'pool group', csv_writer_dict_list, vs_ref)
            # Get the skipepd list for http policy.
            if 'http_policies' in virtual_service:
                csv_object = self.get_csv_object_list(
                    csv_writer_dict_list,
                    ['add cs policy', 'add responder policy',
                     'add rewrite policy'])
                for http_ref in virtual_service['http_policies']:
                    http_name = self.get_name(http_ref['http_policy_set_ref'])
                    skipped_list = self.get_csv_skipped_list(csv_object,
                                                             http_name,
                                                             vs_ref)
                    if skipped_list:
                        skipped_setting['Httppolicy'] = {}
                        skipped_setting['Httppolicy']['name'] = http_name
                        skipped_setting['Httppolicy'][
                            'skipped_list'] = skipped_list
                    # Get the http policy name
                    for each_http_policy in avi_config['HTTPPolicySet']:
                        if each_http_policy['name'] == http_name:
                            for http_req in \
                                    each_http_policy['http_request_policy'][
                                        'rules']:
                                if http_req.get('switching_action', None) and \
                                        http_req['switching_action'].get(
                                            'pool_group_ref', None):
                                    pool_group_name = self.get_name(
                                        http_req['switching_action']
                                        ['pool_group_ref'])
                                    self.get_pool_skipped_list(
                                        avi_config, pool_group_name,
                                        skipped_setting, csv_object,
                                        'Httppolicy',
                                        csv_writer_dict_list, vs_ref)
            # Get the skipped list for application_profile_ref.
            if 'application_profile_ref' in virtual_service and \
                            'admin:System' not in \
                            virtual_service['application_profile_ref']:
                name, skipped = self.get_application_profile_skipped(
                    csv_writer_dict_list, virtual_service, vs_ref)
                if skipped:
                    skipped_setting['Application profile'] = {}
                    skipped_setting['Application profile'][
                        'name'] = name
                    skipped_setting['Application profile'][
                        'skipped_list'] = skipped
            # Get the skipped list for network profile ref.
            if 'network_profile_ref' in virtual_service and \
                            'admin:System' not in \
                            virtual_service['network_profile_ref']:
                name, skipped = self.get_network_profile_skipped(
                    csv_writer_dict_list, virtual_service, vs_ref)
                if skipped:
                    skipped_setting['Network profile'] = {}
                    skipped_setting['Network profile'][
                        'name'] = name
                    skipped_setting['Network profile'][
                        'skipped_list'] = skipped
            # Update overall skipped setting of VS csv row
            if skipped_setting:
                vs_csv_object.update(
                    {'Overall skipped settings': str(skipped_setting)})
            else:
                vs_csv_object.update(
                    {'Overall skipped settings': "FULLY MIGRATION"})
                fully_migrated += 1
            msg = "Writing excel sheet started..."
            self.print_progress_bar(progressbar_count, total_count, msg,
                                    prefix='Progress', suffix='')
        csv_objects = [row for row in csv_writer_dict_list
                       if row['Status'] in [STATUS_PARTIAL, STATUS_SUCCESSFUL]
                       and row['Netscaler Command'] not in ['add cs vserver',
                                                            'add lb vserver']
                       and (
                           'VS Reference' not in row or not row[
                               'VS Reference'])]
        # Update the vs reference not in used if objects are not attached to
        # VS directly or indirectly
        for csv_object in csv_objects:
            csv_object['VS Reference'] = STATUS_NOT_IN_USE

    def write_status_report_and_pivot_table_in_xlsx(self, row_list, output_dir,
                                                report_name, vs_level_status):
        """
        This method writes the status and make pivot table in excel sheet
        :param row_list:
        :param output_dir:
        :param report_name:
        :param vs_level_status:
        :return:
        """
        global total_count
        global progressbar_count
        # List of fieldnames for headers
        if vs_level_status:
            fieldnames = ['Line Number', 'Netscaler Command', 'Object Name',
                          'Full Command', 'Status', 'Skipped settings',
                          'Indirect mapping', 'Not Applicable', 'User Ignored',
                          'Overall skipped settings', 'Complexity Level',
                          'VS Reference', 'AVI Object']
        else:
            fieldnames = ['Line Number', 'Netscaler Command', 'Object Name',
                          'Full Command', 'Status', 'Skipped settings',
                          'Indirect mapping', 'Not Applicable', 'User Ignored',
                          'Complexity Level' , 'AVI Object']
        xlsx_report = output_dir + os.path.sep + ("%s-ConversionStatus.xlsx" %
                                                  report_name)
        # xlsx workbook
        status_wb = Workbook(xlsx_report)
        # xlsx worksheet
        status_ws = status_wb.add_worksheet("Status Sheet")
        # Lock the first row of xls report.
        status_ws.freeze_panes(1, 0)
        first_row = 0
        for header in fieldnames:
            col = fieldnames.index(header)
            status_ws.write(first_row, col, header)
        row = 1
        for row_data in row_list:
            progressbar_count += 1
            for _key, _value in row_data.items():
                col = fieldnames.index(_key)
                status_ws.write(row, col, _value)
            msg = "Writing excel sheet started..."
            self.print_progress_bar(progressbar_count, total_count, msg,
                               prefix='Progress', suffix='')
            row += 1
        status_wb.close()
        # create dataframe for row list
        df = pandas.DataFrame(row_list, columns=fieldnames)
        # create pivot table using pandas
        pivot_table = pandas.pivot_table(df,
                                         index=["Status", "Netscaler Command"],
                                         values=[], aggfunc=[len], fill_value=0)
        # create dataframe for pivot table using pandas
        pivot_df = pandas.DataFrame(pivot_table)
        master_book = load_workbook(xlsx_report)
        master_writer = pandas.ExcelWriter(xlsx_report, engine='openpyxl')
        master_writer.book = master_book
        # Add pivot table in Pivot sheet
        pivot_df.to_excel(master_writer, 'Pivot Sheet')
        master_writer.save()

    def update_skip_duplicates(self, obj, obj_list, obj_type,
                               merge_object_mapping, name, ent_type, prefix,
                               syslist):
        """
        This method merge duplicate objects
        :param obj: Source object to find duplicates for
        :param obj_list: List of object to search duplicates in
        :param obj_type: Type of object to add in converted_objs status
        :param converted_objs: Converted avi object or merged object name
        :param name: Name of the object
        :param default_profile_name : Name of root parent default profile
        :return:
        """
        dup_of = None
        merge_object_mapping[obj_type].update({name: name})
        dup_of, old_name = self.check_for_duplicates(obj, obj_list, obj_type,
                                                merge_object_mapping, ent_type,
                                                prefix,
                                                syslist)
        if dup_of:
            LOG.info(
                "Duplicate profiles: %s merged in %s" % (obj['name'], dup_of))
            # Update value of ssl profile with merged profile
            if old_name in merge_object_mapping[obj_type].keys():
                merge_object_mapping[obj_type].update({old_name: dup_of})
            merge_object_mapping[obj_type].update({name: dup_of})
            return True
        return False

    def create_update_vsvip(self, vip, vsvip_config, tenant_ref, cloud_ref,
                            prefix=None):
        """
        This functions defines that create or update VSVIP object.
        :param vip: vip of VS
        :param vsvip_config: List of vs object
        :param tenant_ref: tenant reference
        :param cloud_ref: cloud reference
        :param prefix: prefix for objects
        :return: None
        """

        # Get the exsting vsvip object list if present
        name = vip + '-vsvip'
        # Added prefix for objects
        if prefix:
            name = prefix + '-' + name
        vsvip = [vip_obj for vip_obj in vsvip_config
                 if vip_obj['name'] == name]
        # If VSVIP object not present then create new VSVIP object.
        if not vsvip:
            vsvip_object = {
                "name": name,
                "tenant_ref": tenant_ref,
                "cloud_ref": cloud_ref,
                "vip": [
                    {
                        "vip_id": "0",
                        "ip_address": {
                            "type": "V4",
                            "addr": vip
                        }
                    }
                ],
            }
            vsvip_config.append(vsvip_object)

    def get_redirect_fail_action(self, url):
        """
        This method returns the fail action dict
        :param url: url
        :return:
        """
        parsed = urlparse.urlparse(url)
        redirect_fail_action = {
            'fail_action': {
                'redirect': {
                    'host': parsed.hostname,
                    'protocol': str(parsed.scheme).upper(),
                    'status_code': "HTTP_REDIRECT_STATUS_CODE_302"
                },
                "type": "FAIL_ACTION_HTTP_REDIRECT"
            }
        }
        if parsed.path:
            redirect_fail_action['fail_action']['redirect']['path'] = \
                str(parsed.path).replace('"', '')
        if parsed.query:
            redirect_fail_action['fail_action']['redirect'][
                'query'] = parsed.query

        return redirect_fail_action

    def cleanup_dupof(self, avi_config):
        """
        This method is used to clean up dup_of key from different AVI objects
        :param avi_config:
        :return:
        """
        self.remove_dup_key(avi_config["ApplicationProfile"])
        self.remove_dup_key(avi_config["NetworkProfile"])
        self.remove_dup_key(avi_config["SSLProfile"])
        self.remove_dup_key(avi_config['PKIProfile'])
        self.remove_dup_key(avi_config["ApplicationPersistenceProfile"])
        self.remove_dup_key(avi_config['HealthMonitor'])

    def update_profile_ref(self, ref, avi_obj, merge_obj_list):
        """
        This method is used to update the profile references which was
        attached at the time of creation
        :param ref:
        :param avi_obj:
        :param merge_obj_list:
        :return:
        """
        for obj in avi_obj:
            obj_ref = obj.get(ref)
            tenant_ref = obj.get('tenant_ref')
            if obj_ref:
                name = self.get_name(obj_ref)
                tenant = self.get_name(tenant_ref)
                if name in merge_obj_list:
                    updated_name = merge_obj_list[name]
                    if ref == 'application_persistence_profile_ref':
                        type_cons = OBJECT_TYPE_APPLICATION_PERSISTENCE_PROFILE
                    if ref == 'application_profile_ref':
                        type_cons = OBJECT_TYPE_APPLICATION_PROFILE
                    obj[ref] = self.get_object_ref(updated_name, type_cons,
                                                   tenant)

    def vs_redirect_http_to_https(self, avi_config, sysdict):

        """
        Removes the VS which is redirected to another VS amd update the
        status and avi object for that VS
        :param avi_config: avi configuration after all conversion
        :param sysdict: system configuration
        :return:
        """

        vsrem = {}
        LOG.debug("Check started for redirect from HTTP VS to HTTPS VS with "
                  "no pool")
        for vs in avi_config['VirtualService']:
            if not vs.get('pool_group_ref') and not vs.get(
                    'application_profile_ref') and vs.get('services', []) and \
                    not all([s.get('enable_ssl', True)for s in vs['services']])\
                    and vs.get('http_policies',[]) and vs['http_policies'][
                    0].get('http_policy_set_ref'):
                polname = self.get_name(vs['http_policies'][0][
                                        'http_policy_set_ref'])
                pol = [pl for pl in avi_config['HTTPPolicySet'] if pl['name']
                        == polname]
                if pol and pol[0].get('http_request_policy', {}).get('rules',
                        []) and pol[0]['http_request_policy']['rules'][0].get(
                        'redirect_action'):
                    iplist = [ip['ip_address']['addr'] for ip in vs.get('vip',
                             []) if ip.get('ip_address',{}).get('addr')] or (
                             [vs['ip_address']['addr']] if vs.get(
                             'ip_address',{}).get('addr') else [])
                    if iplist:
                        for nvs in avi_config['VirtualService']:
                            if vs['name'] != nvs['name'] and [ip for ip in
                               iplist if ip in ([nip['ip_address']['addr']
                               for nip in nvs.get('vip', []) if nip.get(
                               'ip_address',{}).get('addr')] or [nvs[
                               'ip_address']['addr'] if nvs.get(
                               'ip_address',{}).get('addr') else []])]:
                                appname = self.get_name(nvs[
                                            'application_profile_ref']) if \
                                            nvs.get('application_profile_ref') \
                                            else None
                                if appname == 'ns-migrate-http':
                                    LOG.debug("%s has redirect to %s, hence "
                                              "removing %s" % (vs['name'],
                                                       nvs['name'], vs['name']))
                                    vsrem[vs['name']] = nvs['name']
                                appprof = [pr for pr in (avi_config[
                                          'ApplicationProfile'] + sysdict[
                                          'ApplicationProfile']) if pr['name']
                                           == appname]
                                if appprof and appprof[0]['type'] == \
                                        'APPLICATION_PROFILE_TYPE_HTTP':
                                    if appprof[0].get('http_profile'):
                                        appprof[0]['http_profile'][
                                            'http_to_https'] = True
                                    else:
                                        appprof[0]['http_profile'] = {
                                            'http_to_https': True}
                                    LOG.debug("%s has redirect to %s, hence "
                                              "setting 'http_to_https' as true "
                                              "and removing %s" %(vs['name'],
                                                    nvs['name'], vs['name']))
                                    vsrem[vs['name']] = nvs['name']
                                # Condition to merge http ports to https vs
                                if [True for ssl in nvs['services'] if ssl[
                                    'enable_ssl']] and \
                                        [True for ssl_vs in vs['services']
                                         if not ssl_vs['enable_ssl']]:
                                    nvs['services'].append(vs['services'][0])
                                    vsrem[vs['name']] = nvs['name']

        LOG.debug("Check completed for redirect from HTTP VS to HTTPS VS with "
                  "no pool")
        if vsrem:
            avi_config['VirtualService'] = [v for v in avi_config[
                                           'VirtualService'] if v['name'] not
                                            in vsrem.keys()]
            LOG.debug('%s VS got removed from AVI configuration' % str(len(
                        vsrem)))
            for cl in csv_writer_dict_list:
                if cl['Object Name'] in vsrem.keys() and cl[
                   'Netscaler Command'] in ['add lb vserver', 'add cs vserver']:
                    cl['Status'] = STATUS_INDIRECT
                    cl['AVI Object'] = 'Redirected to %s' % vsrem[cl[
                                        'Object Name']]

    def merge_pool(self, avi_config):
        """
        This method merge the pools in AVI if HM is same
        :param avi_config:
        :return:
        """
        mergelist=[]
        for poolgrp in avi_config['PoolGroup']:
            pool_member = poolgrp['members']
            length = len(pool_member)
            for count in range(length):
                pool_name = pool_member[count]['pool_ref'].split('&')[1].split(
                    '=')[1]
                if pool_name in mergelist:
                    continue
                pool = [pl for pl in avi_config['Pool']
                        if pl['name'] == pool_name]
                if not pool:
                    LOG.debug("'%s' not present" % pool_name)
                    continue
                for count2 in range(count+1, length):
                    pname = pool_member[count2]['pool_ref'].split('&')[1].split(
                        '=')[1]
                    nextpool = [pol for pol in avi_config['Pool']
                        if pol['name'] == pname]

                    if not nextpool:
                        LOG.debug("'%s' not present" % pname)
                        continue
                    if pool[0]['health_monitor_refs'].sort() == nextpool[0][
                      'health_monitor_refs'].sort():
                        LOG.debug("Merging pool '%s' in '%s'" % (nextpool[0][
                                                    'name'], pool[0]['name']))
                        ip_port = set()
                        for ser in pool[0]['servers']:
                            ip_port.add(str(ser['ip']['addr']) + ':' + str(
                                ser['port']))
                        for server in nextpool[0]['servers']:
                            ipport = str(server['ip']['addr']) + ':' + str(
                                        server['port'])
                            if ipport not in list(ip_port):
                                pool[0]['servers'].append(server)
                        for cl in csv_writer_dict_list:
                            if cl['Object Name'] == (nextpool[0][
                             'name'].replace('-pool','')) and cl[
                             'Netscaler Command'] in ['add service',
                             'add serviceGroup']:
                                cl['AVI Object'] = 'Merged to %s' % pool[0][
                                                                        'name']
                        mergelist.append(nextpool[0]['name'])
        for plg in avi_config['PoolGroup']:
            plg['members'] = [member for member in plg['members'] if
                              member['pool_ref'].split('&')[1].split('=')[1] not
                              in mergelist]
        avi_config['Pool'] = [pools for pools in avi_config['Pool'] if pools[
                                'name'] not in mergelist]

    def add_policy(self, policy, updated_vs_name, avi_config, tmp_policy_ref,
                   vs_obj, tenant_name, cloud_name, prefix, used_poolgrp_ref):
        """
        This method is used to add policy objects to AVI and also add
        reference in VS
        :param policy: policy object
        :param updated_vs_name: vs name
        :param avi_config: avi config dict
        :param tmp_policy_ref: list of policy ref which are already used
        :param vs_obj: vs object
        :param tenant_name: name of tenant
        :param cloud_name: name of cloud
        :param prefix: prefix
        :param used_poolgrp_ref: list of used pool group ref
        :return:
        """
        if policy['name'] in tmp_policy_ref:
            # clone the http policy set if it is referenced to other VS
            policy = self.clone_http_policy_set(policy, updated_vs_name,
                         avi_config, tenant_name, cloud_name, used_poolgrp_ref,
                         userprefix=prefix)
        updated_http_policy_ref = self.get_object_ref(policy['name'],
                                   OBJECT_TYPE_HTTP_POLICY_SET, tenant_name)

        tmp_policy_ref.append(policy['name'])
        http_policies = {
            'index': 11,
            'http_policy_set_ref': updated_http_policy_ref
        }
        if not vs_obj.get('http_policies'):
            vs_obj['http_policies'] = []
        else:
            ind = max([policies['index'] for policies in vs_obj[
                       'http_policies']])
            http_policies['index'] = ind + 1
        vs_obj['http_policies'].append(http_policies)
        avi_config['HTTPPolicySet'].append(policy)

    def build_redirect_action_dict(self, redirect_url, enable_ssl):
        """
        This method returns a redirect action dict
        :param redirect_url: redirect url
        :param enable_ssl: flag for ssl enable
        :return:
        """
        redirect_url = self.parse_url(redirect_url)
        protocol = str(redirect_url.scheme).upper()
        hostname = str(redirect_url.hostname)
        pathstring = str(redirect_url.path)
        querystring = str(redirect_url.query)
        full_path = '%s?%s' % (pathstring, querystring) if pathstring and \
                                querystring else pathstring
        protocol = enable_ssl and 'HTTPS' or 'HTTP' if not protocol else \
            protocol
        action = {
            'protocol': protocol
        }
        if hostname:
            action.update({'host':
                {
                    'type': 'URI_PARAM_TYPE_TOKENIZED',
                    'tokens': [{
                        'type': 'URI_TOKEN_TYPE_STRING',
                        'str_value': hostname,
                        'start_index': '0',
                        'end_index': '65535'
                    }]
                }
            })
        if full_path:
            action.update({'path':
                {
                    'type': 'URI_PARAM_TYPE_TOKENIZED',
                    'tokens': [{
                        'type': 'URI_TOKEN_TYPE_STRING',
                        'str_value': full_path,
                        'start_index': '0',
                        'end_index': '65535'
                    }]
                }
            })
        return action

    def create_http_to_https_custom_profile(self):
        '''

        :return: custom application profile dict
        '''
        return {
            'name': "ns-migrate-http",
            'type': "APPLICATION_PROFILE_TYPE_HTTP",
            'tenant_ref': "/api/tenant/?name=admin",
            'preserve_client_ip': False,
            'http_profile': {
                'max_rps_uri': 0,
                'keepalive_header': False,
                'max_rps_cip_uri': 0,
                'x_forwarded_proto_enabled': False,
                'connection_multiplexing_enabled': True,
                'websockets_enabled': True,
                'enable_request_body_buffering': False,
                'hsts_enabled': False,
                'xff_enabled': True,
                'disable_keepalive_posts_msie6': True,
                'keepalive_timeout': 30000,
                'ssl_client_certificate_mode': "SSL_CLIENT_CERTIFICATE_NONE",
                'http_to_https': True,
                'spdy_enabled': False,
                'max_bad_rps_cip_uri': 0,
                'client_body_timeout': 30000,
                'httponly_enabled': False,
                'hsts_max_age': 365,
                'max_bad_rps_cip': 0,
                'server_side_redirect_to_https': False,
                'client_max_header_size': 12,
                'client_max_request_size': 48,
                'max_rps_unknown_uri': 0,
                'ssl_everywhere_enabled': False,
                'spdy_fwd_proxy_mode': False,
                'post_accept_timeout': 30000,
                'client_header_timeout': 10000,
                'secure_cookie_enabled': False,
                'xff_alternate_name': "X-Forwarded-For",
                'max_rps_cip': 0,
                'client_max_body_size': 0,
                'max_rps_unknown_cip': 0,
                'allow_dots_in_header_name': False,
                'max_bad_rps_uri': 0,
                'use_app_keepalive_timeout': False
            },
            'dos_rl_profile': {
                'rl_profile': {
                    'client_ip_connections_rate_limit': {
                        'count': 0,
                        'explicit_tracking': False,
                        'period': 1,
                        'action': {
                            'status_code': "HTTP_LOCAL_RESPONSE_STATUS_CODE_429",
                            'type': "RL_ACTION_NONE"
                        },
                        'burst_sz': 0,
                        'fine_grain': False
                    }
                },
                'dos_profile': {
                    'thresh_period': 5
                }
            }
        }

    def correct_vs_ref(self, avi_config):
        """
        This method corrects the reference of VS to different objects
        :param avi_config: avi configuration dict
        :return:
        """
        global csv_writer_dict_list
        avi_graph = self.make_graph(avi_config)
        csv_dict_sub = [row for row in csv_writer_dict_list if row[
            'Netscaler Command'] not in ('add lb vserver',
                                         'add cs vserver') and row[
                            'Status'] in (STATUS_PARTIAL,
                                          STATUS_SUCCESSFUL)]
        for dict_row in csv_dict_sub:
            obj = dict_row['AVI Object']
            if isinstance(obj, str) and obj.startswith('{'):
                vs = []
                if '__/__' in obj:
                    for dataobj in obj.split('__/__'):
                        obj = eval(dataobj)
                        self.add_vs_ref(obj, avi_graph, vs)
                else:
                    obj = eval(obj)
                    self.add_vs_ref(obj, avi_graph, vs)
                if vs:
                    dict_row['VS Reference'] = str(list(set(vs)))
                else:
                    dict_row['VS Reference'] = STATUS_NOT_IN_USE

    def add_vs_ref(self, obj, avi_graph, vs):
        """
        Helper method for adding vs ref
        :param obj: object
        :param avi_graph: avi graph
        :param vs: VS list
        :return:
        """
        obj_name = obj.get('name', obj.get('hostname'))
        if obj_name:
            if avi_graph.has_node(obj_name):
                LOG.debug("Checked predecessor for %s", obj_name)
                predecessor = list(avi_graph.predecessors(obj_name))
                if predecessor:
                    self.get_predecessor(predecessor, avi_graph, vs)
            else:
                LOG.debug("Object %s may be merged or orphaned", obj_name)

    def get_predecessor(self, predecessor, avi_graph, vs):
        """
        This method gets the predecessor of the object
        :param predecessor: predecessor list
        :param avi_graph: avi graph
        :param vs: VS list
        :return:
        """
        if len(predecessor) > 1:
            for node in predecessor:
                nodelist = [node]
                self.get_predecessor(nodelist, avi_graph, vs)
        elif len(predecessor):
            node_obj = [nod for nod in list(avi_graph.nodes().data()) if
                        nod[0] == predecessor[0]]
            if node_obj and (node_obj[0][1]['type'] == 'VS' or 'VS' in node_obj[
              0][1]['type']):
                LOG.debug("Predecessor %s found", predecessor[0])
                vs.extend(predecessor)
            else:
                LOG.debug("Checked predecessor for %s", predecessor[0])
                nodelist = list(avi_graph.predecessors(predecessor[0]))
                self.get_predecessor(nodelist, avi_graph, vs)
        else:
            LOG.debug("No more predecessor")
